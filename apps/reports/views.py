from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, DecimalField, ExpressionWrapper, Case, When, Q
from django.db.models.functions import Coalesce, Cast
from django.http import JsonResponse
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, date
import logging
import csv
import pandas as pd
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from apps.company.models import FiscalYear
from apps.billing.models import Invoice, InvoiceItem, DebitNote, CreditNote
from apps.bookkeeping.models import LedgerAccount, JournalEntry, JournalEntryLine, LedgerOpeningBalance
from apps.reports.services.analytics import (
    get_fiscal_year_trends, get_top_customers, get_top_products,
    get_vendor_stats, get_yoy_sales, get_cash_flow, get_ar_aging,
)
from apps.payments.models import Payment
from apps.products.models import Product, StockTransaction
from apps.customers.models import Customer
from apps.purchasing.models import PurchaseOrder
from apps.billing.models import VendorBill
from apps.payments.models import VendorPayment
from apps.vendors.models import Vendor
from apps.reports.models import Report, UserReportAccess
from apps.reports.report_registry import (
    REPORT_REGISTRY, get_user_visible_reports, get_dashboard_sections,
)
from decimal import Decimal
from apps.utils.nepali_date import bs_str_to_ad, ad_date_to_bs_str

logger = logging.getLogger(__name__)

# Helper function for fiscal year filtering


def filter_by_fiscal_year(queryset, fiscal_year, date_field='created_at'):
    """Filter queryset by fiscal year dates"""
    if fiscal_year:
        filter_kwargs = {
            f'{date_field}__gte': fiscal_year.start_date,
            f'{date_field}__lte': fiscal_year.end_date,
        }
        return queryset.filter(**filter_kwargs)
    return queryset


def get_report_branch(request, company):
    """Return a valid branch for the current user company if branch accounting is enabled."""
    if not company or not getattr(company, 'enable_branch_accounting', False):
        return None

    branch_id = request.GET.get('branch')
    if not branch_id:
        return None

    try:
        return company.branches.filter(pk=int(branch_id)).first()
    except (ValueError, TypeError):
        return None


def apply_branch_filters(queryset, branch, lookup_fields):
    """Apply branch filters across one or more related lookups."""
    if not branch:
        return queryset

    if isinstance(lookup_fields, str):
        lookup_fields = [lookup_fields]

    branch_filter = Q()
    for lookup in lookup_fields:
        branch_filter |= Q(**{lookup: branch})

    return queryset.filter(branch_filter)


def report_branch_context(company, branch):
    return {
        'branches': company.branches.order_by('name') if company and getattr(company, 'enable_branch_accounting', False) else [],
        'selected_branch': branch,
    }


@login_required
@login_required
def report_dashboard(request):
    company = request.user_company
    visible = get_user_visible_reports(request.user, company)
    sections = get_dashboard_sections(visible)
    return render(request, 'reports/report_dashboard.html', {
        'sections': sections,
        'is_admin': request.user.is_superuser or request.user.is_company_admin,
    })


@login_required
def cash_flow_report(request):
    """
    IAS 7 / NFRS-equivalent Cash Flow Statement.

    Direct method — traces every movement through Cash and Bank accounts and
    classifies each counter-account as:
      Operating  — REVENUE or EXPENSE account on the other side
      Investing  — non-current ASSET account on the other side
      Financing  — LIABILITY or EQUITY account on the other side
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()

    cash_bank_accounts = LedgerAccount.objects.filter(
        company=user_company,
        account_type='ASSET',
        name__in=['Cash', 'Bank'],
        is_deleted=False,
    )

    lines_qs = JournalEntryLine.objects.filter(
        account__in=cash_bank_accounts,
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
    ).select_related('journal_entry', 'account')
    lines_qs = filter_by_fiscal_year(lines_qs, fiscal_year, date_field='journal_entry__date')

    # Opening cash balance (before period start)
    opening_cash = Decimal('0.00')
    if fiscal_year:
        open_qs = JournalEntryLine.objects.filter(
            account__in=cash_bank_accounts,
            journal_entry__company=user_company,
            journal_entry__is_deleted=False,
            journal_entry__date__lt=fiscal_year.start_date,
        )
        d = open_qs.aggregate(
            dr=Coalesce(Sum('amount', filter=Q(entry_type='DEBIT')),  Decimal('0')),
            cr=Coalesce(Sum('amount', filter=Q(entry_type='CREDIT')), Decimal('0')),
        )
        opening_cash = d['dr'] - d['cr']

    operating_in  = Decimal('0.00')
    operating_out = Decimal('0.00')
    investing_in  = Decimal('0.00')
    investing_out = Decimal('0.00')
    financing_in  = Decimal('0.00')
    financing_out = Decimal('0.00')

    operating_lines  = []
    investing_lines  = []
    financing_lines  = []

    for line in lines_qs:
        je = line.journal_entry
        # Find counter-account(s) in the same entry
        counter_lines = JournalEntryLine.objects.filter(
            journal_entry=je
        ).exclude(
            account__in=cash_bank_accounts
        ).select_related('account')

        counter_type = None
        counter_is_current = True
        counter_name = je.description or ''
        for cl in counter_lines:
            counter_type = cl.account.account_type
            counter_is_current = cl.account.is_current
            counter_name = cl.account.name
            break  # use first counter-account to classify

        amount = line.amount
        is_inflow = (line.entry_type == 'DEBIT')

        if counter_type in ('REVENUE', 'EXPENSE') or counter_type is None:
            category = 'operating'
        elif counter_type == 'ASSET' and not counter_is_current:
            category = 'investing'
        elif counter_type in ('LIABILITY', 'EQUITY'):
            category = 'financing'
        else:
            category = 'operating'

        row = {
            'date': je.date,
            'description': je.description,
            'account': counter_name,
            'amount': amount,
            'is_inflow': is_inflow,
        }

        if category == 'operating':
            operating_lines.append(row)
            if is_inflow:
                operating_in += amount
            else:
                operating_out += amount
        elif category == 'investing':
            investing_lines.append(row)
            if is_inflow:
                investing_in += amount
            else:
                investing_out += amount
        else:
            financing_lines.append(row)
            if is_inflow:
                financing_in += amount
            else:
                financing_out += amount

    net_operating  = operating_in  - operating_out
    net_investing  = investing_in  - investing_out
    net_financing  = financing_in  - financing_out
    net_cash_flow  = net_operating + net_investing + net_financing
    closing_cash   = opening_cash  + net_cash_flow

    context = {
        'fiscal_year':      fiscal_year,
        'company':          user_company,
        'opening_cash':     opening_cash,
        'closing_cash':     closing_cash,
        'operating_lines':  operating_lines,
        'investing_lines':  investing_lines,
        'financing_lines':  financing_lines,
        'operating_in':     operating_in,
        'operating_out':    operating_out,
        'net_operating':    net_operating,
        'investing_in':     investing_in,
        'investing_out':    investing_out,
        'net_investing':    net_investing,
        'financing_in':     financing_in,
        'financing_out':    financing_out,
        'net_financing':    net_financing,
        'net_cash_flow':    net_cash_flow,
        'report_title':     'Statement of Cash Flows',
    }

    return render(request, 'reports/cash_flow_report.html', context)


@login_required
def sales_summary_report(request):
    """Generate sales summary report"""
    try:
        # noinspection PyUnresolvedReferences
        company = request.user.company
        if not company:
            return JsonResponse({'error': 'No company associated with your account'}, status=400)

        # Get date range from request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        fiscal_year_id = request.GET.get('fiscal_year')

        # Get fiscal year
        fiscal_year = None
        if fiscal_year_id:
            fiscal_year = get_object_or_404(
                FiscalYear, id=fiscal_year_id, company=company)
        else:
            fiscal_year = FiscalYear.objects.filter(
                company=company,
                start_date__lte=timezone.now(),
                end_date__gte=timezone.now()
            ).first()

        if not fiscal_year:
            return JsonResponse({'error': 'No fiscal year found'}, status=400)

        # Get sales data
        sales_qs = Invoice.objects.filter(
            company=company,
            status='ISSUED',
            outstanding_balance=0,
            created_at__range=[fiscal_year.start_date, fiscal_year.end_date]
        )
        branch = get_report_branch(request, company)
        sales_qs = apply_branch_filters(sales_qs, branch, 'branch')

        sales_data = sales_qs.aggregate(
            total_sales=Coalesce(Sum('total'), 0),
            number_of_sales=Count('id'),
            average_sale=Coalesce(
                ExpressionWrapper(
                    Sum('total') / Cast(Count('id'), DecimalField()),
                    output_field=DecimalField()
                ),
                0
            )
        )

        return JsonResponse({
            'total_sales': float(sales_data['total_sales']),
            'number_of_sales': sales_data['number_of_sales'],
            'average_sale': float(sales_data['average_sale'])
        })

    except Exception as e:
        logger.error(f"Error in sales summary report: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def product_sales_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')
    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    product_sales_qs = InvoiceItem.objects.filter(
        product__isnull=False, invoice__company=user_company)
    branch = get_report_branch(request, user_company)
    product_sales_qs = apply_branch_filters(product_sales_qs, branch, 'invoice__branch')
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
    product_sales_qs = filter_by_fiscal_year(
        product_sales_qs, fiscal_year, date_field='invoice__created_at')
    product_sales = product_sales_qs.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('product__name')
    context = {
        'product_sales': product_sales,
        'fiscal_year': fiscal_year,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/product_sales_report.html', context)


@login_required
def service_sales_report(request):
    """Displays a report of sales by service."""
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')
    # Query InvoiceItem for items that are services
    # Filter where description is not null, and product and package are null
    # Group by description and credit account
    # Annotate with total quantity sold and total revenue
    branch = get_report_branch(request, user_company)
    service_sales = InvoiceItem.objects.filter(
        invoice__company=user_company,
        description__isnull=False,
        product__isnull=True,
        package__isnull=True
    )
    service_sales = apply_branch_filters(service_sales, branch, 'invoice__branch')
    service_sales = service_sales.values(
        'description'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('description')

    # You can add date filtering here later if needed

    context = {
        'service_sales': service_sales,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/service_sales_report.html', context)


@login_required
def sales_by_customer_report(request):
    """Displays a report of sales by customer."""
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    # Query Invoice, group by customer, annotate total sales and number of invoices
    sales_by_customer = Invoice.objects.filter(customer__isnull=False, company=user_company)
    sales_by_customer = apply_branch_filters(sales_by_customer, branch, 'branch')
    sales_by_customer = sales_by_customer.values(
        'customer__name'
    ).annotate(
        total_sales_amount=Sum('total'),
        number_of_sales=Count('id')
    ).order_by('customer__name')

    # You can add date filtering here later if needed

    context = {
        'sales_by_customer': sales_by_customer,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/sales_by_customer_report.html', context)


@login_required
def revenue_by_time_report(request):
    """Displays a report of revenue by month and year."""
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    # Query Invoice, group by year and month of created_at
    # Annotate with the total revenue for that period
    revenue_by_time = Invoice.objects.filter(company=user_company)
    revenue_by_time = apply_branch_filters(revenue_by_time, branch, 'branch')
    revenue_by_time = revenue_by_time.values(
        'created_at__year', 'created_at__month'
    ).annotate(
        total_revenue=Sum('total')
    ).order_by('-created_at__year', '-created_at__month')

    # You can add date filtering here later if needed

    context = {
        'revenue_by_time': revenue_by_time,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/revenue_by_time_report.html', context)


@login_required
def detailed_sales_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    branch = get_report_branch(request, user_company)
    invoices = Invoice.active_objects.filter(company=user_company)
    invoices = apply_branch_filters(invoices, branch, 'branch').order_by('-due_date')

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        invoices = invoices.filter(date__gte=start_date)

    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        invoices = invoices.filter(date__lte=end_date)

    total_sales = invoices.aggregate(Sum('total'))['total__sum'] or 0
    total_items_sold = InvoiceItem.objects.filter(invoice__in=invoices).aggregate(Sum('quantity'))[
        'quantity__sum'] or 0

    context = {
        'invoices': invoices,
        'total_sales': total_sales,
        'total_items_sold': total_items_sold,
        'start_date': start_date_str,
        'end_date': end_date_str,
        **report_branch_context(user_company, branch),
    }

    return render(request, 'reports/detailed_sales_report.html', context)


@login_required
def product_performance_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    # Aggregate product sales data
    product_performance = InvoiceItem.objects.filter(invoice__company=user_company,)
    product_performance = apply_branch_filters(product_performance, branch, 'invoice__branch').values(
        'product__name'
    ).annotate(
        total_quantity_sold=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'product_performance': product_performance,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/product_performance_report.html', context)


@login_required
def sales_by_user_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    sales_by_user = Invoice.active_objects.filter(company=user_company)
    sales_by_user = apply_branch_filters(sales_by_user, branch, 'branch').values(
        'created_by'
    ).annotate(
        total_sales=Sum('total')
    ).order_by('-total_sales')

    context = {
        'sales_by_user': sales_by_user,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/sales_by_user_report.html', context)


@login_required
def referral_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    from apps.pos.models import POSSale, Referrer

    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    ref_id = request.GET.get('referrer', '').strip()

    branch = get_report_branch(request, user_company)

    sales_qs = POSSale.active_objects.filter(company=user_company, referred_by__isnull=False)
    sales_qs = apply_branch_filters(sales_qs, branch, 'branch')
    if start_date:
        sales_qs = sales_qs.filter(created_at__date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(created_at__date__lte=end_date)
    if ref_id:
        sales_qs = sales_qs.filter(referred_by__id=ref_id)

    from django.db.models import F
    pos_referral_data = sales_qs.values(
        'referred_by__id', 'referred_by__name', 'referred_by__phone'
    ).annotate(
        sale_count=Count('id'),
        total_sales=Sum('total'),
        total_outstanding=Sum('invoice__outstanding_balance'),
    ).annotate(
        total_collected=ExpressionWrapper(
            Sum('total') - Sum('invoice__outstanding_balance'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).order_by('-total_sales')

    referrers = Referrer.objects.filter(company=user_company).order_by('name')

    context = {
        'pos_referral_data': pos_referral_data,
        'referrers': referrers,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'selected_referrer': ref_id,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/referral_report.html', context)


@login_required
def outstanding_invoices_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    outstanding_invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer')
    outstanding_invoices = apply_branch_filters(outstanding_invoices, branch, 'branch').order_by('-created_at')
    total_outstanding_amount = outstanding_invoices.aggregate(Sum('total'))[
        'total__sum'] or 0.00

    context = {
        'outstanding_invoices': outstanding_invoices,
        'total_outstanding_amount': total_outstanding_amount,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/outstanding_invoices_report.html', context)


@login_required
def payment_history_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    payments = Payment.objects.filter(company=user_company)
    payments = apply_branch_filters(payments, branch, ['branch', 'invoice__branch']).order_by('-date')
    total_payments_received = payments.aggregate(
        Sum('amount'))['amount__sum'] or 0.00

    context = {
        'payments': payments,
        'total_payments_received': total_payments_received,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/payment_history_report.html', context)


@login_required
def debit_note_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    debit_notes = DebitNote.objects.filter(company=user_company)
    debit_notes = apply_branch_filters(debit_notes, branch, ['branch', 'invoice__branch']).order_by('-created_at')
    total_debit_note_amount = debit_notes.aggregate(
        Sum('amount'))['amount__sum'] or 0.00

    context = {
        'debit_notes': debit_notes,
        'total_debit_note_amount': total_debit_note_amount,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/debit_note_report.html', context)


@login_required
def ar_aging_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    today = date.today()
    branch = get_report_branch(request, user_company)
    invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer')
    invoices = apply_branch_filters(invoices, branch, 'branch')

    ar_aging_data = []
    for invoice in invoices:
        if not invoice.due_date:
            continue
        days_past_due = (today - invoice.due_date).days
        balance = invoice.outstanding_balance
        row = {
            'invoice_id': invoice.pk,
            'invoice_number': invoice.invoice_number,
            'customer_name': invoice.customer.name if invoice.customer else '',
            'due_date': invoice.due_date,
            'total_amount': invoice.total,
            'outstanding_balance': balance,
            'current': balance if days_past_due < 0 else 0,
            'days_1_30': balance if 0 <= days_past_due <= 30 else 0,
            'days_31_60': balance if 31 <= days_past_due <= 60 else 0,
            'days_61_90': balance if 61 <= days_past_due <= 90 else 0,
            'days_over_90': balance if days_past_due > 90 else 0,
        }
        ar_aging_data.append(row)

    context = {
        'ar_aging_data': ar_aging_data,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/ar_aging_report.html', context)


@login_required
def sales_by_category_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    branch = get_report_branch(request, user_company)
    sales_by_category = InvoiceItem.objects.filter(
        invoice__company=user_company,
        product__category__isnull=False
    )
    sales_by_category = apply_branch_filters(sales_by_category, branch, 'invoice__branch').values(
        'product__category__name'
    ).annotate(
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'sales_by_category': sales_by_category,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/sales_by_category_report.html', context)


@login_required
def cogs_report(request):
    # noinspection PyUnresolvedReferences
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_product = request.GET.get('product')
    selected_category = request.GET.get('category')

    branch = get_report_branch(request, user_company)
    # Calculate COGS (Cost of Goods Sold)
    # This is a simplified example. A real COGS calculation might involve more complex inventory valuation methods (e.g., FIFO, LIFO, Weighted Average).
    # Here, we assume a direct link between invoice items and product costs.
    cogs_data = InvoiceItem.objects.filter(
        invoice__company=user_company,
        invoice__outstanding_balance=0,  # Filter for paid invoices
        product__isnull=False
    )
    cogs_data = apply_branch_filters(cogs_data, branch, 'invoice__branch').annotate(
        item_cogs=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=DecimalField()
        )
    ).select_related('invoice', 'product', 'product__category')

    # Apply date filters
    if start_date:
        cogs_data = cogs_data.filter(invoice__created_at__date__gte=start_date)
    if end_date:
        cogs_data = cogs_data.filter(invoice__created_at__date__lte=end_date)
    
    # Apply product filter
    if selected_product:
        cogs_data = cogs_data.filter(product__id=selected_product)
    
    # Apply category filter
    if selected_category:
        cogs_data = cogs_data.filter(product__category__id=selected_category)

    cogs_data = cogs_data.order_by('invoice__created_at')

    total_cogs = cogs_data.aggregate(
        total=Sum('item_cogs')
    )['total'] or 0.00

    # Get all products and categories for filter dropdowns
    from apps.products.models import Product, Category
    all_products = Product.objects.filter(company=user_company).order_by('name')
    all_categories = Category.objects.filter(company=user_company).order_by('name')

    context = {
        'cogs_data': cogs_data,
        'total_cogs': total_cogs,
        'all_products': all_products,
        'all_categories': all_categories,
        'selected_start_date': start_date,
        'selected_end_date': end_date,
        'selected_product': selected_product,
        'selected_category': selected_category,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/billing/cogs_report.html', context)


@login_required
def profitability_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    # This is a simplified profitability report (Gross Profit)
    # Gross Profit = Total Sales - Total COGS
    # More advanced reports would include operating expenses, taxes, etc.

    branch = get_report_branch(request, user_company)
    # Calculate Total Sales
    total_sales = Invoice.objects.filter(company=user_company)
    total_sales = apply_branch_filters(total_sales, branch, 'branch').aggregate(
        Sum('total'))['total__sum'] or Decimal('0.00')

    # Calculate Total COGS (re-using logic from cogs_report)
    cogs_data = InvoiceItem.objects.filter(
        invoice__company=user_company,
        invoice__outstanding_balance=0,  # Filter for paid invoices
        product__isnull=False
    )
    cogs_data = apply_branch_filters(cogs_data, branch, 'invoice__branch').annotate(
        line_item_cogs=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=DecimalField()
        )
    )
    total_cogs = cogs_data.aggregate(Sum('line_item_cogs'))[
        'line_item_cogs__sum'] or Decimal('0.00')

    gross_profit = total_sales - total_cogs

    context = {
        'total_sales': total_sales,
        'total_cogs': total_cogs,
        'gross_profit': gross_profit,
        **report_branch_context(user_company, branch),
    }
    return render(request, 'reports/profitability_report.html', context)


@login_required
def inventory_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    from django.core.paginator import Paginator

    products = Product.objects.filter(company=user_company).order_by('name')
    low_stock_products = products.filter(
        productstock__stock__lte=F('productstock__minimum_stock'))

    paginator = Paginator(products, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'products': page_obj,
        'products_with_stock': page_obj,
        'low_stock_products': low_stock_products,
        'page_obj': page_obj,
    }
    return render(request, 'reports/products/inventory_report.html', context)


@login_required
def stock_movement_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    qs = StockTransaction.objects.filter(
        product__company=user_company
    ).select_related('product', 'user').order_by('-created_at')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    product_q = request.GET.get('product', '').strip()
    txn_type = request.GET.get('txn_type', '')
    stock_type = request.GET.get('stock_type', '')

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if product_q:
        qs = qs.filter(product__name__icontains=product_q)
    if txn_type:
        qs = qs.filter(transaction_type=txn_type)
    if stock_type:
        qs = qs.filter(stock_type=stock_type)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'stock_movements': page_obj,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'product_q': product_q,
        'txn_type': txn_type,
        'stock_type': stock_type,
        'total_count': paginator.count,
    }
    return render(request, 'reports/products/stock_movement_report.html', context)


@login_required
def customer_acquisition_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    customers = Customer.objects.filter(
        company=user_company).order_by('-created_at')
    new_customers_this_month = customers.filter(created_at__month=timezone.now(
    ).month, created_at__year=timezone.now().year).count()
    total_customers = customers.count()

    context = {
        'customers': customers,
        'new_customers_this_month': new_customers_this_month,
        'total_customers': total_customers,
    }
    return render(request, 'reports/customers/customer_acquisition_report.html', context)


@login_required
def purchase_order_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    qs = PurchaseOrder.active_objects.filter(
        company=user_company
    ).select_related('vendor').order_by('-date')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', '')
    vendor_q = request.GET.get('vendor', '').strip()

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if status:
        qs = qs.filter(status=status)
    if vendor_q:
        qs = qs.filter(vendor__name__icontains=vendor_q)

    total_po_amount = qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'purchase_orders': page_obj,
        'page_obj': page_obj,
        'total_po_amount': total_po_amount,
        'total_count': paginator.count,
        'company': user_company,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'vendor_q': vendor_q,
        'PURCHASE_STATUS_CHOICES': [('DRAFT','Draft'),('SENT','Sent'),('RECEIVED','Received'),('CANCELLED','Cancelled')],
    }
    return render(request, 'reports/purchasing/purchase_order_list_report.html', context)


@login_required
def vendor_bill_list_report(request):
    if request.user.is_superuser:
        qs = VendorBill.objects.all()
    else:
        user_company = request.user_company
        if not user_company:
            messages.warning(request, "Your account is not associated with a company. Please contact an administrator.")
            return redirect('accounts:user_dashboard')
        qs = VendorBill.objects.filter(vendor__company=user_company)

    qs = qs.select_related('vendor', 'purchase_order').prefetch_related('payments').order_by('-bill_date')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', '')
    vendor_q = request.GET.get('vendor', '').strip()

    if date_from:
        qs = qs.filter(bill_date__gte=date_from)
    if date_to:
        qs = qs.filter(bill_date__lte=date_to)
    if status:
        qs = qs.filter(status=status)
    if vendor_q:
        qs = qs.filter(vendor__name__icontains=vendor_q)

    from django.db.models import OuterRef, Subquery
    from apps.payments.models import VendorPayment as VP
    paid_subq = VP.objects.filter(vendor_bill=OuterRef('pk')).values('vendor_bill').annotate(s=Sum('amount')).values('s')
    qs = qs.annotate(paid_amount=Subquery(paid_subq))
    total_billed = qs.aggregate(t=Sum('total_amount'))['t'] or 0

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'vendor_bills': page_obj,
        'page_obj': page_obj,
        'total_billed': total_billed,
        'total_count': paginator.count,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'vendor_q': vendor_q,
        'VENDOR_BILL_STATUS_CHOICES': [('UNPAID','Unpaid'),('PARTIAL','Partial'),('PAID','Paid'),('CANCELLED','Cancelled')],
    }
    return render(request, 'reports/purchasing/vendor_bill_list_report.html', context)


@login_required
def vendor_payment_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    qs = VendorPayment.objects.filter(
        vendor_bill__vendor__company=user_company
    ).select_related('vendor_bill__vendor', 'bank_account').order_by('-payment_date')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    vendor_q = request.GET.get('vendor', '').strip()
    method = request.GET.get('method', '')

    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)
    if vendor_q:
        qs = qs.filter(vendor_bill__vendor__name__icontains=vendor_q)
    if method:
        qs = qs.filter(payment_method=method)

    total_payments = qs.aggregate(Sum('amount'))['amount__sum'] or 0
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'vendor_payments': page_obj,
        'page_obj': page_obj,
        'total_payments': total_payments,
        'total_count': paginator.count,
        'date_from': date_from,
        'date_to': date_to,
        'vendor_q': vendor_q,
        'method': method,
    }
    return render(request, 'reports/purchasing/vendor_payment_list_report.html', context)


@login_required
def ledger_account_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    accounts = LedgerAccount.objects.filter(
        company=user_company, is_deleted=False
    ).order_by('name')

    context = {
        'accounts': accounts,
        'ledger_accounts': accounts,  # alias for backward compat
    }
    return render(request, 'reports/bookkeeping/ledger_account_list_report.html', context)


@login_required
def payment_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.active_objects.filter(
        company=user_company).order_by('-date')

    context = {
        'payments': payments
    }
    return render(request, 'reports/bookkeeping/payment_list_report.html', context)


@login_required
def journal_entry_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')
    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    journal_entries = JournalEntry.objects.filter(
        company=user_company).order_by('-date')
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
    journal_entries = filter_by_fiscal_year(
        journal_entries, fiscal_year, date_field='date')
    # Search/filtering
    description = request.GET.get('description', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if description:
        journal_entries = journal_entries.filter(
            description__icontains=description)
    if date_from:
        journal_entries = journal_entries.filter(date__gte=date_from)
    if date_to:
        journal_entries = journal_entries.filter(date__lte=date_to)
    paginator = Paginator(journal_entries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'entries': page_obj,           # template variable name
        'journal_entries': page_obj,   # alias
        'fiscal_year': fiscal_year,
        'page_obj': page_obj,
        'search': {
            'description': description,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'reports/bookkeeping/journal_entry_list_report.html', context)


@login_required
def export_detailed_sales_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    invoices = Invoice.objects.filter(
        company=user_company).order_by('-due_date')

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        invoices = invoices.filter(date__gte=start_date)

    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        invoices = invoices.filter(date__lte=end_date)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="detailed_sales_report.xlsx"'

    df = pd.DataFrame(list(invoices.values('invoice_number',
                      'customer__name', 'total', 'created_at')))
    df.rename(columns={
        'invoice_number': 'Invoice Number',
        'customer__name': 'Customer Name',
        'total': 'Total Amount',
        'created_at': 'Date',
    }, inplace=True)

    # Convert datetime objects to timezone-aware datetime objects, then localize them
    # Then convert to a simple date string or just date object for Excel compatibility
    if 'Date' in df.columns:
        # Remove timezone info for simpler excel export
        df['Date'] = df['Date'].dt.tz_localize(None)

    df.to_excel(response, index=False, sheet_name='Detailed Sales Report')

    return response


@login_required
def export_detailed_sales_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    invoices = Invoice.objects.filter(
        company=user_company).order_by('-due_date')

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        invoices = invoices.filter(date__gte=start_date)

    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        invoices = invoices.filter(due_date__lte=end_date)

    total_sales = invoices.aggregate(Sum('total'))['total__sum'] or 0.00
    total_items_sold = InvoiceItem.objects.filter(
        invoice__in=invoices).aggregate(Sum('quantity'))['quantity__sum'] or 0

    context = {
        'invoices': invoices,
        'total_sales': total_sales,
        'total_items_sold': total_items_sold,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/pdf/detailed_sales_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:detailed_sales_report')


@login_required
def print_detailed_sales_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    invoices = Invoice.objects.filter(
        company=user_company).order_by('-due_date')

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        invoices = invoices.filter(date__gte=start_date)

    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        invoices = invoices.filter(date__lte=end_date)

    total_sales = invoices.aggregate(Sum('total'))['total__sum'] or 0.00
    total_items_sold = InvoiceItem.objects.filter(
        invoice__in=invoices).aggregate(Sum('quantity'))['quantity__sum'] or 0

    context = {
        'invoices': invoices,
        'total_sales': total_sales,
        'total_items_sold': total_items_sold,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'company': user_company,
    }

    return render(request, 'reports/print/detailed_sales_report_print.html', context)


@login_required
def export_product_performance_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    product_performance = InvoiceItem.objects.filter(invoice__company=user_company).values(
        'product__name'
    ).annotate(
        total_quantity_sold=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="product_performance_report.xlsx"'

    df = pd.DataFrame(list(product_performance))
    df.rename(columns={
        'product__name': 'Product Name',
        'total_quantity_sold': 'Total Quantity Sold',
        'total_revenue': 'Total Revenue',
    }, inplace=True)

    df.to_excel(response, index=False, sheet_name='Product Performance Report')

    return response


@login_required
def export_product_performance_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    product_performance = InvoiceItem.objects.filter(invoice__company=user_company).values(
        'product__name'
    ).annotate(
        total_quantity_sold=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'product_performance': product_performance,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/pdf/product_performance_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:product_performance_report')


@login_required
def print_product_performance_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    product_performance = InvoiceItem.objects.filter(invoice__company=user_company).values(
        'product__name'
    ).annotate(
        total_quantity_sold=Sum('quantity'),
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'product_performance': product_performance,
        'company': user_company,
    }

    return render(request, 'reports/print/product_performance_report_print.html', context)


@login_required
def export_sales_by_user_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_user = Invoice.objects.filter(company=user_company
                                           ).values(
        'created_by__username'
    ).annotate(
        total_sales=Sum('total')
    ).order_by('-total_sales')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_by_user_report.xlsx"'

    df = pd.DataFrame(list(sales_by_user))
    df.rename(columns={
        'salesperson__email': 'Salesperson Email',
        'salesperson__first_name': 'First Name',
        'salesperson__last_name': 'Last Name',
        'total_sales_amount': 'Total Sales Amount',
        'number_of_invoices': 'Number of Invoices',
    }, inplace=True)

    df.to_excel(response, index=False, sheet_name='Sales By User Report')

    return response


@login_required
def export_sales_by_user_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_user = Invoice.objects.filter(company=user_company).values(
        'created_by__username'
    ).annotate(
        total_sales=Sum('total')
    ).order_by('-total_sales')

    context = {
        'sales_by_user': sales_by_user,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/pdf/sales_by_user_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:sales_by_user_report')


@login_required
def print_sales_by_user_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_user = Invoice.objects.filter(company=user_company).values(
        'created_by__username'
    ).annotate(
        total_sales=Sum('total')
    ).order_by('-total_sales')

    context = {
        'sales_by_user': sales_by_user,
        'company': user_company,
    }

    return render(request, 'reports/print/sales_by_user_report_print.html', context)


@login_required
def view_report(request, report_id):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    # Ensure report belongs to user's company
    report = get_object_or_404(
        Report, id=report_id, created_by__company=user_company)
    data = {}

    if report.report_type == 'sales':
        invoices = Invoice.objects.filter(company=user_company, created_at__date__gte=report.start_date,
                                          created_at__date__lte=report.end_date)
        total_sales = invoices.aggregate(total=Sum('total'))['total'] or 0
        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices).values(
            'product__name', 'package__name', 'description', 'price').annotate(total_quantity=Sum('quantity'),
                                                                               total_amount=Sum(ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())))
        data = {
            'invoices': invoices,
            'total_sales': total_sales,
            'invoice_items': invoice_items
        }
    elif report.report_type == 'inventory':
        products = Product.objects.filter(company=user_company)
        low_stock_products = Product.objects.filter(
            company=user_company, productstock__stock__lte=F('productstock__minimum_stock'))
        data = {
            'products': products,
            'low_stock_products': low_stock_products
        }
    elif report.report_type == 'stock_transactions':
        transactions = StockTransaction.objects.filter(
            product__company=user_company,  # Filter by product's company
            created_at__date__gte=report.start_date,
            created_at__date__lte=report.end_date
        ).order_by('-created_at')
        data = {'transactions': transactions}
    elif report.report_type == 'billing':
        invoices = Invoice.objects.filter(company=user_company, created_at__date__gte=report.start_date,
                                          created_at__date__lte=report.end_date)
        total_invoices = invoices.aggregate(total=Sum('total'))['total'] or 0
        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices).values(
            'product__name', 'package__name', 'price').annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum(ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())))
        data = {
            'invoices': invoices,
            'total_invoices': total_invoices,
            'invoice_items': invoice_items
        }

    return render(request, 'reports/view_report.html', {
        'report': report,
        'data': data
    })


@login_required
def export_report_csv(request, report_id):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    report = get_object_or_404(
        Report, id=report_id, created_by__company=user_company)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report.name}_{report.report_type}.csv"'

    writer = csv.writer(response)

    if report.report_type == 'sales':
        writer.writerow(['Invoice ID', 'Date', 'User', 'Total'])
        invoices = Invoice.objects.filter(company=user_company, created_at__date__gte=report.start_date,
                                          created_at__date__lte=report.end_date)
        for invoice in invoices:
            writer.writerow([invoice.id, invoice.created_at,
                            invoice.user.email, invoice.total])
        writer.writerow([])
        writer.writerow(['Item/Package/Description', 'Price',
                        'Total Quantity', 'Total Amount'])
        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices).values(
            'product__name', 'package__name', 'description', 'price').annotate(total_quantity=Sum('quantity'),
                                                                               total_amount=Sum(ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())))
        for item in invoice_items:
            # Handle item name based on type
            name = "Unknown Item"
            if item['product__name']:
                name = item['product__name']
            elif item['package__name']:
                name = f"Package: {item['package__name']}"
            elif item['description']:
                name = item['description']

            writer.writerow([name, item['price'],
                             item['total_quantity'], item['total_amount']])

    elif report.report_type == 'inventory':
        writer.writerow(['Item', 'Barcode', 'Category',
                        'Stock', 'Price', 'Minimum Stock'])
        products = Product.objects.filter(company=user_company).select_related('category', 'productstock')
        for product in products:
            writer.writerow([
                product.name, product.barcode, product.category.name,
                product.productstock.stock if hasattr(
                    product, 'productstock') else 0, product.price,
                product.productstock.minimum_stock if hasattr(
                    product, 'productstock') else 0
            ])

    elif report.report_type == 'stock_transactions':
        writer.writerow(['Date', 'Item', 'Type', 'Quantity', 'Reason', 'User'])
        transactions = StockTransaction.objects.filter(
            product__company=user_company,
            created_at__date__gte=report.start_date,
            created_at__date__lte=report.end_date
        ).select_related('product', 'user')
        for transaction in transactions:
            writer.writerow([transaction.created_at, transaction.product.name,
                             transaction.get_transaction_type_display(), transaction.quantity,
                             transaction.reason, transaction.user.email])

    elif report.report_type == 'billing':
        writer.writerow(['Invoice ID', 'Date', 'User', 'Total'])
        invoices = Invoice.objects.filter(
            company=user_company,
            created_at__date__gte=report.start_date,
            created_at__date__lte=report.end_date
        )
        for invoice in invoices:
            writer.writerow([invoice.id, invoice.created_at,
                            invoice.user.email, invoice.total])
        writer.writerow([])
        writer.writerow(
            ['Item/Package', 'Price', 'Total Quantity', 'Total Amount'])
        invoice_items = InvoiceItem.objects.filter(invoice__in=invoices).values(
            'product__name', 'package__name', 'price').annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum(ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())))
        for item in invoice_items:
            name = item['product__name'] if item['product__name'] else item['package__name']
            writer.writerow(
                [name, item['price'], item['total_quantity'], item['total_amount']])

    return response


@login_required
def export_report_pdf(request, report_id):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    report = get_object_or_404(
        Report, id=report_id, created_by__company=user_company)
    start_date = None
    end_date = None
    fiscal_year = None
    revenue_accounts = []
    expense_accounts = []
    total_revenue = 0
    total_expenses = 0
    net_profit_loss = 0

    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if start_date_str:
            converted = bs_str_to_ad(start_date_str)
            if converted is not None:
                start_date = converted
        if end_date_str:
            converted = bs_str_to_ad(end_date_str)
            if converted is not None:
                end_date = converted
    else:
        # Use fiscal year if no date range is provided
        fiscal_year_id = request.session.get('active_fiscal_year_id')
        if fiscal_year_id:
            fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
            if fiscal_year:
                start_date = fiscal_year.start_date
                end_date = fiscal_year.end_date

    # Get Revenue and Expense accounts for the user's company
    all_revenue_accounts = LedgerAccount.objects.filter(
        account_type='REVENUE', company=user_company)
    all_expense_accounts = LedgerAccount.objects.filter(
        account_type='EXPENSE', company=user_company)

    date_min = start_date if start_date else datetime.min.date()
    date_max = end_date if end_date else datetime.max.date()

    # Single grouped query replaces N per-account aggregation queries
    pl_totals = (
        JournalEntryLine.objects
        .filter(
            account__company=user_company,
            account__account_type__in=['REVENUE', 'EXPENSE'],
            journal_entry__date__gte=date_min,
            journal_entry__date__lte=date_max,
        )
        .values('account_id', 'account__name', 'account__account_type', 'entry_type')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0.00')))
    )

    # Build lookup: {account_id: {entry_type: total}}
    pl_lookup: dict = {}
    for row in pl_totals:
        pl_lookup.setdefault(row['account_id'], {
            'name': row['account__name'],
            'type': row['account__account_type'],
        }).update({row['entry_type']: row['total']})

    account_objects = {a.id: a for a in list(all_revenue_accounts) + list(all_expense_accounts)}

    for account_id, data in pl_lookup.items():
        acc = account_objects.get(account_id)
        if acc is None:
            continue
        credit = data.get('CREDIT', Decimal('0.00'))
        debit = data.get('DEBIT', Decimal('0.00'))
        if data['type'] == 'REVENUE':
            net = credit - debit
            revenue_accounts.append({'account': acc, 'total': net})
            total_revenue += net
        else:
            net = debit - credit
            expense_accounts.append({'account': acc, 'total': net})
            total_expenses += net

    # Calculate Net Profit/Loss
    net_profit_loss = total_revenue - total_expenses

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'start_date_bs': ad_date_to_bs_str(start_date) if start_date else '',
        'end_date_bs': ad_date_to_bs_str(end_date) if end_date else '',
        'revenue_accounts': revenue_accounts,
        'expense_accounts': expense_accounts,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit_loss': net_profit_loss,
        'fiscal_year': fiscal_year,
    }
    return render(request, 'reports/profit_and_loss_report.html', context)


@login_required
def export_profit_and_loss_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Fetching data for Profit and Loss Report
    # Revenue (Credit entries to Revenue accounts)
    revenue_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='REVENUE')
    revenue_entries = JournalEntryLine.objects.filter(
        account__in=revenue_accounts,
        journal_entry__company=user_company,
        entry_type='CREDIT'
    )

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        revenue_entries = revenue_entries.filter(
            journal_entry__date__gte=start_date)
    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        revenue_entries = revenue_entries.filter(
            journal_entry__date__lte=end_date)

    total_revenue = revenue_entries.aggregate(
        Sum('amount'))['amount__sum'] or Decimal('0.00')

    # Expenses (Debit entries to Expense accounts)
    expense_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EXPENSE')
    expense_entries = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__company=user_company,
        entry_type='DEBIT'
    )

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        expense_entries = expense_entries.filter(
            journal_entry__date__gte=start_date)
    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        expense_entries = expense_entries.filter(
            journal_entry__date__lte=end_date)

    total_expenses = expense_entries.aggregate(
        Sum('amount'))['amount__sum'] or Decimal('0.00')

    net_profit = total_revenue - total_expenses

    # Create a DataFrame for Excel export
    data = {
        'Category': ['Total Revenue', 'Total Expenses', 'Net Profit'],
        'Amount': [total_revenue, total_expenses, net_profit]
    }
    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="profit_and_loss_report.xlsx"'

    df.to_excel(response, index=False, sheet_name='Profit and Loss')

    return response


@login_required
def export_profit_and_loss_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    revenue_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='REVENUE')
    revenue_entries = JournalEntryLine.objects.filter(
        account__in=revenue_accounts,
        journal_entry__company=user_company,
        entry_type='CREDIT'
    )

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        revenue_entries = revenue_entries.filter(
            journal_entry__date__gte=start_date)
    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        revenue_entries = revenue_entries.filter(
            journal_entry__date__lte=end_date)

    total_revenue = revenue_entries.aggregate(
        Sum('amount'))['amount__sum'] or Decimal('0.00')

    expense_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EXPENSE')
    expense_entries = JournalEntryLine.objects.filter(
        account__in=expense_accounts,
        journal_entry__company=user_company,
        entry_type='DEBIT'
    )

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
        expense_entries = expense_entries.filter(
            journal_entry__date__gte=start_date)
    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)
        expense_entries = expense_entries.filter(
            journal_entry__date__lte=end_date)

    total_expenses = expense_entries.aggregate(
        Sum('amount'))['amount__sum'] or Decimal('0.00')

    net_profit = total_revenue - total_expenses

    context = {
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/print_profit_and_loss_report.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:profit_and_loss_report')


@login_required
def print_profit_and_loss_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    start_date = bs_str_to_ad(start_date_str) if start_date_str else None
    end_date = bs_str_to_ad(end_date_str) if end_date_str else None

    rev_qs = LedgerAccount.objects.filter(company=user_company, account_type='REVENUE')
    revenue_entries = JournalEntryLine.objects.filter(
        account__in=rev_qs, journal_entry__company=user_company, entry_type='CREDIT'
    )
    if start_date:
        revenue_entries = revenue_entries.filter(journal_entry__date__gte=start_date)
    if end_date:
        revenue_entries = revenue_entries.filter(journal_entry__date__lte=end_date)

    exp_qs = LedgerAccount.objects.filter(company=user_company, account_type='EXPENSE')
    expense_entries = JournalEntryLine.objects.filter(
        account__in=exp_qs, journal_entry__company=user_company, entry_type='DEBIT'
    )
    if start_date:
        expense_entries = expense_entries.filter(journal_entry__date__gte=start_date)
    if end_date:
        expense_entries = expense_entries.filter(journal_entry__date__lte=end_date)

    from apps.utils.nepali_date import ad_date_to_bs_str
    from django.utils import timezone as tz

    rev_totals = (
        revenue_entries.values('account__id', 'account__name')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0.00')))
    )
    exp_totals = (
        expense_entries.values('account__id', 'account__name')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0.00')))
    )

    rev_account_map = {a.id: a for a in rev_qs}
    exp_account_map = {a.id: a for a in exp_qs}

    revenue_accounts_data = [
        {'account': rev_account_map[r['account__id']], 'total': r['total']}
        for r in rev_totals if r['account__id'] in rev_account_map
    ]
    expense_accounts_data = [
        {'account': exp_account_map[e['account__id']], 'total': e['total']}
        for e in exp_totals if e['account__id'] in exp_account_map
    ]

    total_revenue = sum(r['total'] for r in revenue_accounts_data) or Decimal('0.00')
    total_expenses = sum(e['total'] for e in expense_accounts_data) or Decimal('0.00')
    net_profit = total_revenue - total_expenses

    context = {
        'revenue_accounts': revenue_accounts_data,
        'expense_accounts': expense_accounts_data,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'company': user_company,
        'start_date_bs': start_date_str or '',
        'end_date_bs': end_date_str or '',
        'today': tz.now().date(),
    }
    return render(request, 'reports/print_profit_and_loss_report.html', context)


@login_required
def debit_credit_note_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    credit_notes = CreditNote.objects.filter(
        company=user_company).order_by('-created_at')
    debit_notes = DebitNote.objects.filter(
        company=user_company).order_by('-created_at')

    # You could add filtering logic here based on request.GET or a form

    context = {
        'credit_notes': credit_notes,
        'debit_notes': debit_notes,
    }
    return render(request, 'reports/debit_credit_note_report.html', context)


@login_required
def trial_balance_report(request):
    """Generate trial balance report showing all ledger accounts with their debit and credit balances."""
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    report_date = timezone.now().date()  # Default to today
    
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        if fiscal_year:
            report_date = fiscal_year.end_date

    # Handle date filtering from form
    if request.method == 'POST':
        report_date_str = request.POST.get('report_date')
        if report_date_str:
            try:
                converted = bs_str_to_ad(report_date_str)
                if converted is None:
                    raise ValueError("Unparseable date")
                report_date = converted
            except ValueError:
                messages.error(request, "Invalid date format.")
                report_date = timezone.now().date()

    # Get all ledger accounts for the company
    ledger_accounts = LedgerAccount.objects.filter(company=user_company).order_by('account_type', 'name')
    
    trial_balance_data = []
    total_debits = Decimal('0.00')
    total_credits = Decimal('0.00')

    # Single grouped query replaces 2×N per-account queries
    tb_rows = (
        JournalEntryLine.objects
        .filter(
            account__company=user_company,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
        )
        .values('account_id', 'entry_type')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0.00')))
    )
    tb_lookup: dict = {}
    for row in tb_rows:
        tb_lookup.setdefault(row['account_id'], {})[row['entry_type']] = row['total']

    account_map = {a.id: a for a in ledger_accounts}

    for account in ledger_accounts:
        sums = tb_lookup.get(account.id, {})
        debit_balance = sums.get('DEBIT', Decimal('0.00'))
        credit_balance = sums.get('CREDIT', Decimal('0.00'))

        if account.account_type in ['ASSET', 'EXPENSE']:
            net_balance = debit_balance - credit_balance
            if net_balance > 0:
                trial_balance_data.append({
                    'account': account,
                    'debit_balance': net_balance,
                    'credit_balance': Decimal('0.00'),
                })
                total_debits += net_balance
            elif net_balance < 0:
                trial_balance_data.append({
                    'account': account,
                    'debit_balance': Decimal('0.00'),
                    'credit_balance': abs(net_balance),
                })
                total_credits += abs(net_balance)
        else:
            net_balance = credit_balance - debit_balance
            if net_balance > 0:
                trial_balance_data.append({
                    'account': account,
                    'debit_balance': Decimal('0.00'),
                    'credit_balance': net_balance,
                })
                total_credits += net_balance
            elif net_balance < 0:
                trial_balance_data.append({
                    'account': account,
                    'debit_balance': abs(net_balance),
                    'credit_balance': Decimal('0.00'),
                })
                total_debits += abs(net_balance)
    
    # Filter out accounts with zero balance
    trial_balance_data = [item for item in trial_balance_data if item['debit_balance'] > 0 or item['credit_balance'] > 0]
    
    # Check if trial balance is balanced
    is_balanced = abs(total_debits - total_credits) < Decimal('0.01')
    difference = total_debits - total_credits
    
    context = {
        'report_date': report_date,
        'report_date_bs': ad_date_to_bs_str(report_date),
        'trial_balance_data': trial_balance_data,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'difference': difference,
        'is_balanced': is_balanced,
        'fiscal_year': fiscal_year,
        'company': user_company,
        'report_title': 'Trial Balance Report',
    }

    return render(request, 'reports/trial_balance_report.html', context)


@login_required
def balance_sheet_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    report_date = timezone.now().date()  # Default to today
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        if fiscal_year:
            report_date = fiscal_year.end_date

    if request.method == 'POST':
        report_date_str = request.POST.get('report_date')
        if report_date_str:
            try:
                converted = bs_str_to_ad(report_date_str)
                if converted is None:
                    raise ValueError("Unparseable date")
                report_date = converted
            except ValueError:
                messages.error(request, "Invalid date format.")
                report_date = timezone.now().date()

    # Calculate balances for all relevant accounts up to the report_date, filtered by company
    assets = []
    liabilities = []
    equity = []
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    relevant_accounts = LedgerAccount.objects.filter(
        account_type__in=['ASSET', 'LIABILITY', 'EQUITY'], company=user_company)

    # Single grouped query replaces 2×N per-account queries
    bs_rows = (
        JournalEntryLine.objects
        .filter(
            account__company=user_company,
            account__account_type__in=['ASSET', 'LIABILITY', 'EQUITY'],
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
        )
        .values('account_id', 'entry_type')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0.00')))
    )
    bs_lookup: dict = {}
    for row in bs_rows:
        bs_lookup.setdefault(row['account_id'], {})[row['entry_type']] = row['total']

    for account in relevant_accounts:
        sums = bs_lookup.get(account.id, {})
        balance_debits = sums.get('DEBIT', Decimal('0.00'))
        balance_credits = sums.get('CREDIT', Decimal('0.00'))
        if account.account_type == 'ASSET':
            balance = balance_debits - balance_credits
            assets.append({'account': account, 'balance': balance})
            total_assets += balance
        elif account.account_type == 'LIABILITY':
            balance = balance_credits - balance_debits
            liabilities.append({'account': account, 'balance': balance})
            total_liabilities += balance
        elif account.account_type == 'EQUITY':
            balance = balance_credits - balance_debits
            equity.append({'account': account, 'balance': balance})
            total_equity += balance
    # Use filter_by_fiscal_year for net income calculation
    revenue_qs = JournalEntryLine.objects.filter(
        account__account_type='REVENUE',
        journal_entry__company=user_company,
        entry_type='CREDIT'
    )
    expenses_qs = JournalEntryLine.objects.filter(
        account__account_type='EXPENSE',
        journal_entry__company=user_company,
        entry_type='DEBIT'
    )
    if fiscal_year:
        revenue_qs = filter_by_fiscal_year(
            revenue_qs, fiscal_year, date_field='journal_entry__date')
        expenses_qs = filter_by_fiscal_year(
            expenses_qs, fiscal_year, date_field='journal_entry__date')
    else:
        revenue_qs = revenue_qs.filter(journal_entry__date__lte=report_date)
        expenses_qs = expenses_qs.filter(journal_entry__date__lte=report_date)
    revenue_up_to_date = revenue_qs.aggregate(
        sum_amount=Coalesce(Sum('amount'), Decimal('0.00')))['sum_amount']
    expenses_up_to_date = expenses_qs.aggregate(
        sum_amount=Coalesce(Sum('amount'), Decimal('0.00')))['sum_amount']
    net_income_up_to_date = revenue_up_to_date - expenses_up_to_date

    current_assets          = [a for a in assets      if a['account'].is_current]
    noncurrent_assets       = [a for a in assets      if not a['account'].is_current]
    current_liabilities     = [l for l in liabilities if l['account'].is_current]
    noncurrent_liabilities  = [l for l in liabilities if not l['account'].is_current]
    total_equity_all = total_equity + net_income_up_to_date

    context = {
        'report_date': report_date,
        'report_date_bs': ad_date_to_bs_str(report_date),
        'assets': assets,
        'current_assets': current_assets,
        'noncurrent_assets': noncurrent_assets,
        'liabilities': liabilities,
        'current_liabilities': current_liabilities,
        'noncurrent_liabilities': noncurrent_liabilities,
        'equity': equity,
        'current_year_profit': net_income_up_to_date,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity_all,
        'balance_sheet_balanced': abs(total_assets - (total_liabilities + total_equity_all)) < Decimal('0.01'),
        'report_title': 'Statement of Financial Position',
        'fiscal_year': fiscal_year,
        'company': user_company,
    }
    return render(request, 'reports/balance_sheet_report.html', context)


@login_required
def export_balance_sheet_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    report_date_str = request.GET.get('report_date')
    report_date = bs_str_to_ad(report_date_str) if report_date_str else timezone.now().date()

    # Assets
    asset_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='ASSET')
    assets_data = []
    total_assets = 0
    for account in asset_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_debits - balance_credits
        assets_data.append({'account': account, 'balance': balance})
        total_assets += balance

    # Liabilities
    liability_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='LIABILITY')
    liabilities_data = []
    total_liabilities = 0
    for account in liability_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        # Liabilities have natural credit balance
        balance = balance_credits - balance_debits
        liabilities_data.append({'account': account, 'balance': balance})
        total_liabilities += balance

    # Equity (Owners' Equity)
    equity_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EQUITY')
    equity_data = []
    total_equity = 0
    for account in equity_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_credits - balance_debits
        equity_data.append({'account': account, 'balance': balance})
        total_equity += balance

    # Prepare data for Excel export
    data = []
    data.append(['Assets', ''])
    data.extend([[item['account'].name, item['balance']]
                for item in assets_data])
    data.append(['Total Assets', float(total_assets)])
    data.append(['', ''])
    data.append(['Liabilities', ''])
    data.extend([[item['account'].name, item['balance']]
                for item in liabilities_data])
    data.append(['Total Liabilities', float(total_liabilities)])
    data.append(['', ''])
    data.append(['Equity', ''])
    data.extend([[item['account'].name, item['balance']]
                for item in equity_data])
    data.append(['Total Equity', float(total_equity)])
    data.append(['', ''])
    data.append(['Liabilities + Equity',
                float(total_liabilities + total_equity)])

    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="balance_sheet_report.xlsx"'

    df.to_excel(response, index=False, header=False,
                sheet_name='Balance Sheet')

    return response


@login_required
def export_balance_sheet_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    report_date_str = request.GET.get('report_date')
    report_date = bs_str_to_ad(report_date_str) if report_date_str else timezone.now().date()

    asset_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='ASSET')
    assets_data = []
    total_assets = 0
    for account in asset_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_debits - balance_credits
        assets_data.append({'account': account, 'balance': balance})
        total_assets += balance

    liability_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='LIABILITY')
    liabilities_data = []
    total_liabilities = 0
    for account in liability_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_credits - balance_debits
        liabilities_data.append({'account': account, 'balance': balance})
        total_liabilities += balance

    equity_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EQUITY')
    equity_data = []
    total_equity = 0
    for account in equity_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_credits - balance_debits
        equity_data.append({'account': account, 'balance': balance})
        total_equity += balance

    context = {
        'report_date': report_date,
        'assets_data': assets_data,
        'total_assets': total_assets,
        'liabilities_data': liabilities_data,
        'total_liabilities': total_liabilities,
        'equity_data': equity_data,
        'total_equity': total_equity,
        'total_liabilities_and_equity': total_liabilities + total_equity,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/pdf/balance_sheet_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:balance_sheet_report')


@login_required
def print_balance_sheet_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    report_date_str = request.GET.get('report_date')
    report_date = bs_str_to_ad(report_date_str) if report_date_str else timezone.now().date()

    asset_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='ASSET')
    assets_data = []
    total_assets = 0
    for account in asset_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_debits - balance_credits
        assets_data.append({'account': account, 'balance': balance})
        total_assets += balance

    liability_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='LIABILITY')
    liabilities_data = []
    total_liabilities = 0
    for account in liability_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_credits - balance_debits
        liabilities_data.append({'account': account, 'balance': balance})
        total_liabilities += balance

    equity_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EQUITY')
    equity_data = []
    total_equity = 0
    for account in equity_accounts:
        balance_debits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='DEBIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance_credits = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=user_company,
            journal_entry__date__lte=report_date,
            entry_type='CREDIT'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = balance_credits - balance_debits
        equity_data.append({'account': account, 'balance': balance})
        total_equity += balance

    context = {
        'report_date': report_date,
        'assets_data': assets_data,
        'total_assets': total_assets,
        'liabilities_data': liabilities_data,
        'total_liabilities': total_liabilities,
        'equity_data': equity_data,
        'total_equity': total_equity,
        'total_liabilities_and_equity': total_liabilities + total_equity,
        'company': user_company,
    }
    return render(request, 'reports/print/balance_sheet_report_print.html', context)


@login_required
def export_ledger_account_list_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    ledger_accounts = LedgerAccount.objects.filter(
        company=user_company).order_by('name')

    data = []
    data.append(['Account Name', 'Account Type', 'Description'])
    for account in ledger_accounts:
        data.append(
            [account.name, account.get_account_type_display(), account.description])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ledger_account_list.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Ledger Accounts')
    return response


@login_required
def export_ledger_account_list_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    ledger_accounts = LedgerAccount.objects.filter(
        company=user_company).order_by('name')
    context = {
        'ledger_accounts': ledger_accounts,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/ledger_account_list_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:ledger_account_list_report')


@login_required
def print_ledger_account_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    ledger_accounts = LedgerAccount.objects.filter(
        company=user_company).order_by('name')
    context = {
        'ledger_accounts': ledger_accounts,
        'company': user_company,
    }
    return render(request, 'reports/print/ledger_account_list_print.html', context)


@login_required
def export_payment_list_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(
        company=user_company).order_by('-date')

    data = []
    data.append(['Payment Date', 'Amount',
                'Reference Number', 'Payment Method'])
    for payment in payments:
        data.append([payment.date, payment.amount,
                    payment.reference_number, payment.get_method_display()])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_list.xlsx"'
    df.to_excel(response, index=False, header=False, sheet_name='Payments')
    return response


@login_required
def export_payment_list_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(company=user_company)
    context = {
        'payments': payments,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/payment_history_report.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:payment_list_report')


@login_required
def print_payment_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(
        company=user_company).order_by('-date')
    context = {
        'payments': payments,
        'company': user_company,
    }
    return render(request, 'reports/payment_history_report.html', context)


@login_required
def export_journal_entry_list_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    journal_entries = JournalEntry.objects.filter(
        company=user_company).order_by('-date')

    data = []
    data.append(['Date', 'Journal Entry Number', 'Description'])
    for entry in journal_entries:
        data.append(
            [entry.date, entry.journal_entry_number, entry.description])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="journal_entry_list.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Journal Entries')
    return response


@login_required
def export_journal_entry_list_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    journal_entries = JournalEntry.objects.filter(
        company=user_company).order_by('-date')
    context = {
        'journal_entries': journal_entries,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/journal_entry_list_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:journal_entry_list_report')


@login_required
def print_journal_entry_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    journal_entries = JournalEntry.objects.filter(
        company=user_company).order_by('-date')
    context = {
        'journal_entries': journal_entries,
        'company': user_company,
    }
    return render(request, 'reports/print/journal_entry_list_print.html', context)


@login_required
def export_stock_movement_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    qs = StockTransaction.objects.filter(
        product__company=user_company).select_related('product', 'user').order_by('-created_at')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    product_q = request.GET.get('product', '').strip()
    txn_type = request.GET.get('txn_type', '')
    stock_type = request.GET.get('stock_type', '')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if product_q:
        qs = qs.filter(product__name__icontains=product_q)
    if txn_type:
        qs = qs.filter(transaction_type=txn_type)
    if stock_type:
        qs = qs.filter(stock_type=stock_type)

    data = [['Date', 'Product', 'Transaction Type', 'Qty', 'Channel', 'Reason', 'By']]
    for transaction in qs:
        data.append([
            transaction.created_at.strftime('%Y-%m-%d %H:%M'),
            transaction.product.name,
            transaction.get_transaction_type_display(),
            transaction.quantity,
            transaction.stock_type,
            transaction.reason,
            transaction.user.get_full_name() if transaction.user else 'System',
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_movement_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Stock Movement')
    return response


@login_required
def export_stock_movement_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    transactions = StockTransaction.objects.filter(
        product__company=user_company).order_by('-created_at')
    context = {
        'transactions': transactions,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/stock_movement_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:stock_movement_report')


@login_required
def print_stock_movement_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    qs = StockTransaction.objects.filter(
        product__company=user_company).select_related('product', 'user').order_by('-created_at')

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    product_q = request.GET.get('product', '').strip()
    txn_type = request.GET.get('txn_type', '')
    stock_type_filter = request.GET.get('stock_type', '')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if product_q:
        qs = qs.filter(product__name__icontains=product_q)
    if txn_type:
        qs = qs.filter(transaction_type=txn_type)
    if stock_type_filter:
        qs = qs.filter(stock_type=stock_type_filter)

    context = {
        'transactions': qs,
        'company': user_company,
        'report_title': 'Stock Movement Report',
        'today': date.today(),
    }
    return render(request, 'reports/print/stock_movement_report_print.html', context)


@login_required
def export_customer_acquisition_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    customers = Customer.objects.filter(
        company=user_company).order_by('-created_at')

    data = []
    data.append(['Customer Name', 'Email', 'Phone',
                'Address', 'Acquisition Date'])
    for customer in customers:
        data.append([
            customer.name,
            customer.email,
            customer.phone_number,
            customer.address,
            customer.created_at.strftime('%Y-%m-%d')
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customer_acquisition_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Customer Acquisition')
    return response


@login_required
def export_customer_acquisition_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    customers = Customer.objects.filter(
        company=user_company).order_by('-created_at')
    context = {
        'customers': customers,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/customer_acquisition_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:customer_acquisition_report')


@login_required
def print_customer_acquisition_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    customers = Customer.objects.filter(
        company=user_company).order_by('-created_at')
    context = {
        'customers': customers,
        'company': user_company,
        'today': date.today(),
    }
    return render(request, 'reports/print/customer_acquisition_report_print.html', context)


@login_required
def export_outstanding_invoices_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    outstanding_invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer').order_by('-due_date')

    data = []
    data.append(['Invoice Number', 'Customer', 'Total Amount', 'Due Date'])
    for invoice in outstanding_invoices:
        data.append([
            invoice.invoice_number,
            invoice.customer.name if invoice.customer else 'N/A',
            invoice.total,
            invoice.due_date.strftime('%Y-%m-%d')
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="outstanding_invoices_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Outstanding Invoices')
    return response


@login_required
def export_outstanding_invoices_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    outstanding_invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer').order_by('-due_date')
    context = {
        'outstanding_invoices': outstanding_invoices,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/outstanding_invoices_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:outstanding_invoices_report')


@login_required
def print_outstanding_invoices_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    outstanding_invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer').order_by('-due_date')
    context = {
        'outstanding_invoices': outstanding_invoices,
        'company': user_company,
    }
    return render(request, 'reports/print/outstanding_invoices_report_print.html', context)


@login_required
def export_payment_history_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(
        company=user_company).order_by('-date')

    data = []
    data.append(['Payment Date', 'Amount', 'Invoice Number',
                'Customer', 'Payment Method'])
    for payment in payments:
        invoice_number = payment.invoice.invoice_number if payment.invoice else 'N/A'
        customer_name = payment.invoice.customer.name if payment.invoice and payment.invoice.customer else 'N/A'
        data.append([
            payment.date,
            payment.amount,
            invoice_number,
            customer_name,
            payment.get_method_display()
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_history_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Payment History')
    return response


@login_required
def export_payment_history_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(
        company=user_company).order_by('-date')
    context = {
        'payments': payments,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/payment_history_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:payment_history_report')


@login_required
def print_payment_history_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    payments = Payment.objects.filter(
        company=user_company).order_by('-date')
    context = {
        'payments': payments,
        'company': user_company,
    }
    return render(request, 'reports/print/payment_history_report_print.html', context)


@login_required
def export_debit_note_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    debit_notes = DebitNote.objects.filter(
        company=user_company).order_by('-created_at')

    data = []
    data.append(['Debit Note Number', 'Customer', 'Amount', 'Date'])
    for note in debit_notes:
        data.append([
            note.debit_note_number,
            note.customer.name if note.customer else 'N/A',
            note.amount,
            note.date.strftime('%Y-%m-%d')
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="debit_note_report.xlsx"'
    df.to_excel(response, index=False, header=False, sheet_name='Debit Notes')
    return response


@login_required
def export_debit_note_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    debit_notes = DebitNote.objects.filter(
        company=user_company).order_by('-created_at')
    context = {
        'debit_notes': debit_notes,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/debit_note_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:debit_note_report')


@login_required
def print_debit_note_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    debit_notes = DebitNote.objects.filter(
        company=user_company).order_by('-created_at')
    context = {
        'debit_notes': debit_notes,
        'company': user_company,
        'report_title': 'Debit Note Report',
        'today': date.today(),
    }
    return render(request, 'reports/print/debit_note_report_print.html', context)


@login_required
def export_ar_aging_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    today = date.today()
    invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer')

    aging_data = []
    for invoice in invoices:
        if not invoice.due_date:
            continue
        days_past_due = (today - invoice.due_date).days
        aging_category = ''
        if 0 <= days_past_due <= 30:
            aging_category = '0-30 Days'
        elif 31 <= days_past_due <= 60:
            aging_category = '31-60 Days'
        elif 61 <= days_past_due <= 90:
            aging_category = '61-90 Days'
        elif days_past_due > 90:
            aging_category = '>90 Days'

        aging_data.append({
            'invoice_number': invoice.invoice_number,
            'customer_name': invoice.customer.name if invoice.customer else 'N/A',
            'total_amount': invoice.total,
            'due_date': invoice.due_date,
            'days_past_due': days_past_due,
            'aging_category': aging_category
        })

    df = pd.DataFrame(aging_data)

    if not df.empty:
        # Pivot the table to show aging categories as columns
        pivot_table = df.pivot_table(values='total_amount', index='customer_name',
                                     columns='aging_category', aggfunc='sum', fill_value=0)
        # Reorder columns to a standard aging report format
        ordered_columns = ['0-30 Days', '31-60 Days', '61-90 Days', '>90 Days']
        # Ensure all columns exist, add missing ones with zeros
        for col in ordered_columns:
            if col not in pivot_table.columns:
                pivot_table[col] = 0
        pivot_table = pivot_table[ordered_columns]
        pivot_table['Total Outstanding'] = pivot_table.sum(axis=1)
    else:
        pivot_table = pd.DataFrame(
            columns=['0-30 Days', '31-60 Days', '61-90 Days', '>90 Days', 'Total Outstanding'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ar_aging_report.xlsx"'

    pivot_table.to_excel(response, index=True, sheet_name='AR Aging Report')

    return response


@login_required
def export_ar_aging_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    today = date.today()
    invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer')

    aging_30_days = []
    aging_60_days = []
    aging_90_days = []
    aging_over_90_days = []

    for invoice in invoices:
        if not invoice.due_date:
            continue
        days_past_due = (today - invoice.due_date).days
        if 0 <= days_past_due <= 30:
            aging_30_days.append(invoice)
        elif 31 <= days_past_due <= 60:
            aging_60_days.append(invoice)
        elif 61 <= days_past_due <= 90:
            aging_90_days.append(invoice)
        elif days_past_due > 90:
            aging_over_90_days.append(invoice)

    context = {
        'aging_30_days': aging_30_days,
        'aging_60_days': aging_60_days,
        'aging_90_days': aging_90_days,
        'aging_over_90_days': aging_over_90_days,
        'company': user_company,
    }

    html_string = render_to_string(
        'reports/pdf/ar_aging_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:ar_aging_report')


@login_required
def print_ar_aging_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    today = date.today()
    invoices = Invoice.objects.filter(
        company=user_company, outstanding_balance__gt=0).select_related('customer')

    aging_30_days = []
    aging_60_days = []
    aging_90_days = []
    aging_over_90_days = []

    for invoice in invoices:
        if not invoice.due_date:
            continue
        days_past_due = (today - invoice.due_date).days
        if 0 <= days_past_due <= 30:
            aging_30_days.append(invoice)
        elif 31 <= days_past_due <= 60:
            aging_60_days.append(invoice)
        elif 61 <= days_past_due <= 90:
            aging_90_days.append(invoice)
        elif days_past_due > 90:
            aging_over_90_days.append(invoice)

    context = {
        'aging_30_days': aging_30_days,
        'aging_60_days': aging_60_days,
        'aging_90_days': aging_90_days,
        'aging_over_90_days': aging_over_90_days,
        'company': user_company,
        'report_title': 'AR Aging Report',
        'today': today,
    }

    return render(request, 'reports/print/ar_aging_report_print.html', context)


@login_required
def export_sales_by_category_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_category = InvoiceItem.objects.filter(
        invoice__company=user_company,
        product__category__isnull=False
    ).values(
        'product__category__name'
    ).annotate(
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    df = pd.DataFrame(list(sales_by_category))
    df.rename(columns={
        'product__category__name': 'Category',
        'total_revenue': 'Total Revenue',
    }, inplace=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_by_category_report.xlsx"'
    df.to_excel(response, index=False, sheet_name='Sales By Category')
    return response


@login_required
def export_sales_by_category_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_category = InvoiceItem.objects.filter(
        invoice__company=user_company,
        product__category__isnull=False
    ).values(
        'product__category__name'
    ).annotate(
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'sales_by_category': sales_by_category,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/sales_by_category_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:sales_by_category_report')


@login_required
def print_sales_by_category_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    sales_by_category = InvoiceItem.objects.filter(
        invoice__company=user_company,
        product__category__isnull=False
    ).values(
        'product__category__name'
    ).annotate(
        total_revenue=Sum(ExpressionWrapper(
            F('quantity') * F('price'), output_field=DecimalField()))
    ).order_by('-total_revenue')

    context = {
        'sales_by_category': sales_by_category,
        'company': user_company,
    }
    return render(request, 'reports/print/sales_by_category_report_print.html', context)


@login_required
def export_cogs_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    cogs_data = InvoiceItem.objects.filter(
        invoice__company=user_company,
        invoice__outstanding_balance=0,  # Filter for paid invoices
        product__isnull=False
    ).annotate(
        cogs_per_item=F('product__cost_price'),
        line_item_cogs=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=DecimalField()
        )
    ).values(
        'product__name',
        'cogs_per_item',
        'quantity',
        'line_item_cogs'
    ).order_by('product__name')

    data = []
    data.append(['Product Name', 'COGS Per Item',
                'Quantity Sold', 'Line Item COGS'])
    for item in cogs_data:
        data.append([item['product__name'], item['cogs_per_item'],
                    item['quantity'], item['line_item_cogs']])

    total_cogs = sum(item['line_item_cogs'] for item in cogs_data)
    data.append(['Total COGS', '', '', total_cogs])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cogs_report.xlsx"'
    df.to_excel(response, index=False, header=False, sheet_name='COGS Report')
    return response


@login_required
def export_cogs_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    cogs_data = InvoiceItem.objects.filter(
        invoice__company=user_company,
        invoice__outstanding_balance=0,
        product__isnull=False
    ).annotate(
        cogs_per_item=F('product__cost_price'),
        line_item_cogs=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=DecimalField()
        )
    ).values(
        'product__name',
        'cogs_per_item',
        'quantity',
        'line_item_cogs'
    ).order_by('product__name')

    total_cogs = sum(item['line_item_cogs'] for item in cogs_data)

    context = {
        'cogs_data': cogs_data,
        'total_cogs': total_cogs,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/cogs_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:cogs_report')


@login_required
def print_cogs_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    cogs_data = InvoiceItem.objects.filter(
        invoice__company=user_company,
        invoice__outstanding_balance=0,
        product__isnull=False
    ).annotate(
        cogs_per_item=F('product__cost_price'),
        line_item_cogs=ExpressionWrapper(
            F('quantity') * F('product__cost_price'),
            output_field=DecimalField()
        )
    ).values(
        'product__name',
        'cogs_per_item',
        'quantity',
        'line_item_cogs'
    ).order_by('product__name')

    total_cogs = sum(item['line_item_cogs'] for item in cogs_data)

    context = {
        'cogs_data': cogs_data,
        'total_cogs': total_cogs,
        'company': user_company,
    }
    return render(request, 'reports/print/cogs_report_print.html', context)


@login_required
def export_vendor_list_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendors = Vendor.objects.filter(company=user_company).order_by('name')
    context = {
        'vendors': vendors,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/vendor_list_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:vendor_list_report')


@login_required
def print_vendor_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendors = Vendor.objects.filter(company=user_company).order_by('name')
    context = {
        'vendors': vendors,
        'company': user_company,
        'today': date.today(),
    }
    return render(request, 'reports/print/vendor_list_report_print.html', context)


@login_required
def export_purchase_order_list_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    purchase_orders = PurchaseOrder.objects.filter(
        company=user_company).order_by("-date")

    data = []
    data.append(['PO Number', 'Vendor', 'Total Amount',
                'Order Date', 'Status'])
    for po in purchase_orders:
        data.append([
            po.po_number,
            po.vendor.name if po.vendor else 'N/A',
            po.total_amount,
            po.date.strftime('%Y-%m-%d'),
            po.get_status_display()
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="purchase_order_list_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Purchase Orders')
    return response


@login_required
def export_purchase_order_list_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    purchase_orders = PurchaseOrder.objects.filter(
        company=user_company).order_by("-date")
    context = {
        'purchase_orders': purchase_orders,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/purchase_order_list_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:purchase_order_list_report')


@login_required
def print_purchase_order_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    purchase_orders = PurchaseOrder.active_objects.filter(
        company=user_company).order_by("-date")
    context = {
        'purchase_orders': purchase_orders,
        'company': user_company,
    }
    return render(request, 'reports/print/purchase_order_list_report_print.html', context)


@login_required
def export_vendor_bill_list_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_bills = VendorBill.objects.filter(
        company=user_company).order_by('-bill_date')

    data = []
    data.append(['Bill Number', 'Vendor', 'Total Amount',
                'Bill Date', 'Due Date', 'Status'])
    for bill in vendor_bills:
        data.append([
            bill.bill_number,
            bill.vendor.name if bill.vendor else 'N/A',
            bill.total_amount,
            bill.bill_date.strftime('%Y-%m-%d'),
            bill.due_date.strftime('%Y-%m-%d'),
            bill.get_status_display()
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vendor_bill_list_report.xlsx"'
    df.to_excel(response, index=False, header=False, sheet_name='Vendor Bills')
    return response


@login_required
def export_vendor_bill_list_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_bills = VendorBill.objects.filter(
        company=user_company).order_by('-bill_date')
    context = {
        'vendor_bills': vendor_bills,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/vendor_bill_list_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:vendor_bill_list_report')


@login_required
def print_vendor_bill_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_bills = VendorBill.objects.filter(
        company=user_company).order_by('-bill_date')
    context = {
        'vendor_bills': vendor_bills,
        'company': user_company,
    }
    return render(request, 'reports/print/vendor_bill_list_report_print.html', context)


@login_required
def export_vendor_payment_list_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_payments = VendorPayment.objects.filter(
        company=user_company).order_by('-payment_date')

    data = []
    data.append(['Payment Date', 'Vendor', 'Amount', 'Reference'])
    for payment in vendor_payments:
        data.append([
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.vendor.name if payment.vendor else 'N/A',
            payment.amount,
            payment.reference_number
        ])

    df = pd.DataFrame(data)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vendor_payment_list_report.xlsx"'
    df.to_excel(response, index=False, header=False,
                sheet_name='Vendor Payments')
    return response


@login_required
def export_vendor_payment_list_report_pdf(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_payments = VendorPayment.objects.filter(
        company=user_company).order_by('-payment_date')
    context = {
        'vendor_payments': vendor_payments,
        'company': user_company,
    }
    html_string = render_to_string(
        'reports/pdf/vendor_payment_list_report_pdf.html', context)
    messages.error(
        request, "PDF export is not fully configured. WeasyPrint is not installed or configured.")
    return redirect('reports:vendor_payment_list_report')


@login_required
def print_vendor_payment_list_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    vendor_payments = VendorPayment.objects.filter(
        company=user_company).order_by('-payment_date')
    context = {
        'vendor_payments': vendor_payments,
        'company': user_company,
    }
    return render(request, 'reports/print/vendor_payment_list_report_print.html', context)


@login_required
def fiscal_year_dashboard(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')
    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    labels = sales_data = expenses_data = profit_data = []
    top_customers = top_products = []
    yoy_labels = yoy_sales = yoy_prev_sales = []
    cashflow_labels = inflow_data = outflow_data = []
    ar_aging = {}
    vendor_stats = []
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
        if fiscal_year:
            labels, sales_data, expenses_data, profit_data = get_fiscal_year_trends(
                user_company, fiscal_year)
            top_customers = get_top_customers(user_company, fiscal_year)
            top_products = get_top_products(user_company, fiscal_year)
            yoy_labels, yoy_sales, yoy_prev_sales = get_yoy_sales(
                user_company, fiscal_year)
            cashflow_labels, inflow_data, outflow_data = get_cash_flow(
                user_company, fiscal_year)
            ar_aging = get_ar_aging(user_company, fiscal_year)
            vendor_stats = get_vendor_stats(user_company, fiscal_year)
    context = {
        'fiscal_year': fiscal_year,
        'labels': labels,
        'sales_data': sales_data,
        'expenses_data': expenses_data,
        'profit_data': profit_data,
        'top_customers': top_customers,
        'top_products': top_products,
        'yoy_labels': yoy_labels,
        'yoy_sales': yoy_sales,
        'yoy_prev_sales': yoy_prev_sales,
        'cashflow_labels': cashflow_labels,
        'inflow_data': inflow_data,
        'outflow_data': outflow_data,
        'ar_aging': ar_aging,
        'vendor_stats': vendor_stats,
    }
    return render(request, 'reports/fiscal_year_dashboard.html', context)


def export_fiscal_year_dashboard_csv(request):
    user_company = request.user_company
    fiscal_year_id = request.session.get('active_fiscal_year_id')
    if not user_company or not fiscal_year_id:
        return HttpResponse('No company or fiscal year selected.', status=400)
    fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
    if not fiscal_year:
        return HttpResponse('Fiscal year not found.', status=404)
    # Gather analytics data
    labels, sales_data, expenses_data, profit_data = get_fiscal_year_trends(
        user_company, fiscal_year)
    top_customers = get_top_customers(user_company, fiscal_year)
    top_products = get_top_products(user_company, fiscal_year)
    yoy_labels, yoy_sales, yoy_prev_sales = get_yoy_sales(
        user_company, fiscal_year)
    cashflow_labels, inflow_data, outflow_data = get_cash_flow(
        user_company, fiscal_year)
    ar_aging = get_ar_aging(user_company, fiscal_year)
    vendor_stats = get_vendor_stats(user_company, fiscal_year)
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="fiscal_year_dashboard_{fiscal_year.start_date}_{fiscal_year.end_date}.csv"'
    writer = csv.writer(response)
    # Write summary cards
    writer.writerow(['Summary'])
    writer.writerow(['Total Sales', sum(sales_data)])
    writer.writerow(['Total Expenses', sum(expenses_data)])
    writer.writerow(['Total Profit', sum(profit_data)])
    writer.writerow(['Avg. Monthly Profit', sum(profit_data) /
                    len(profit_data) if profit_data else 0])
    writer.writerow([])
    # Write trends
    writer.writerow(['Month', 'Sales', 'Expenses', 'Profit'])
    for i, month in enumerate(labels):
        writer.writerow(
            [month, sales_data[i], expenses_data[i], profit_data[i]])
    writer.writerow([])
    # Cash flow
    writer.writerow(['Month', 'Inflows', 'Outflows'])
    for i, month in enumerate(cashflow_labels):
        writer.writerow([month, inflow_data[i], outflow_data[i]])
    writer.writerow([])
    # AR Aging
    writer.writerow(['A/R Aging'])
    writer.writerow(['Bucket', 'Amount'])
    for bucket, amount in ar_aging.items():
        writer.writerow([bucket, amount])
    writer.writerow([])
    # Top Vendors
    writer.writerow(['Top 5 Vendors'])
    writer.writerow(['Vendor', 'Total Paid'])
    for v in vendor_stats:
        writer.writerow([v['name'], v['total']])
    writer.writerow([])
    # Top Customers
    writer.writerow(['Top 5 Customers'])
    writer.writerow(['Customer', 'Total Sales'])
    for c in top_customers:
        writer.writerow([c['name'], c['total']])
    writer.writerow([])
    # Top Products
    writer.writerow(['Top 5 Products'])
    writer.writerow(['Product', 'Total Sales'])
    for p in top_products:
        writer.writerow([p['name'], p['total']])
    writer.writerow([])
    # YoY Sales
    writer.writerow(['Month', 'Current FY Sales', 'Previous FY Sales'])
    for i, month in enumerate(yoy_labels):
        writer.writerow([month, yoy_sales[i], yoy_prev_sales[i]])
    return response


def export_fiscal_year_dashboard_excel(request):
    user_company = request.user_company
    fiscal_year_id = request.session.get('active_fiscal_year_id')
    if not user_company or not fiscal_year_id:
        return HttpResponse('No company or fiscal year selected.', status=400)
    fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
    if not fiscal_year:
        return HttpResponse('Fiscal year not found.', status=404)
    # Gather analytics data
    labels, sales_data, expenses_data, profit_data = get_fiscal_year_trends(
        user_company, fiscal_year)
    top_customers = get_top_customers(user_company, fiscal_year)
    top_products = get_top_products(user_company, fiscal_year)
    yoy_labels, yoy_sales, yoy_prev_sales = get_yoy_sales(
        user_company, fiscal_year)
    cashflow_labels, inflow_data, outflow_data = get_cash_flow(
        user_company, fiscal_year)
    ar_aging = get_ar_aging(user_company, fiscal_year)
    vendor_stats = get_vendor_stats(user_company, fiscal_year)
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard Analytics'
    row = 1
    # Summary
    ws['A%d' % row] = 'Summary'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Total Sales', sum(sales_data)])
    row += 1
    ws.append(['Total Expenses', sum(expenses_data)])
    row += 1
    ws.append(['Total Profit', sum(profit_data)])
    row += 1
    ws.append(['Avg. Monthly Profit', sum(profit_data) /
              len(profit_data) if profit_data else 0])
    row += 2
    # Trends
    ws['A%d' % row] = 'Trends (Month, Sales, Expenses, Profit)'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Month', 'Sales', 'Expenses', 'Profit'])
    for i, month in enumerate(labels):
        ws.append([month, sales_data[i], expenses_data[i], profit_data[i]])
    row = ws.max_row + 2
    # Cash Flow
    ws['A%d' % row] = 'Cash Flow (Month, Inflows, Outflows)'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Month', 'Inflows', 'Outflows'])
    for i, month in enumerate(cashflow_labels):
        ws.append([month, inflow_data[i], outflow_data[i]])
    row = ws.max_row + 2
    # AR Aging
    ws['A%d' % row] = 'A/R Aging'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Bucket', 'Amount'])
    for bucket, amount in ar_aging.items():
        ws.append([bucket, amount])
    row = ws.max_row + 2
    # Top Vendors
    ws['A%d' % row] = 'Top 5 Vendors'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Vendor', 'Total Paid'])
    for v in vendor_stats:
        ws.append([v['name'], v['total']])
    row = ws.max_row + 2
    # Top Customers
    ws['A%d' % row] = 'Top 5 Customers'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Customer', 'Total Sales'])
    for c in top_customers:
        ws.append([c['name'], c['total']])
    row = ws.max_row + 2
    # Top Products
    ws['A%d' % row] = 'Top 5 Products'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Product', 'Total Sales'])
    for p in top_products:
        ws.append([p['name'], p['total']])
    row = ws.max_row + 2
    # YoY Sales
    ws['A%d' % row] = 'Year-over-Year Sales Comparison'
    ws['A%d' % row].font = Font(bold=True)
    row += 1
    ws.append(['Month', 'Current FY Sales', 'Previous FY Sales'])
    for i, month in enumerate(yoy_labels):
        ws.append([month, yoy_sales[i], yoy_prev_sales[i]])
    # Autosize columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'fiscal_year_dashboard_{fiscal_year.start_date}_{fiscal_year.end_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def profit_and_loss_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(
            request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_date_str = request.GET.get('start_date')
    end_date_str   = request.GET.get('end_date')
    fiscal_year_id = request.session.get('active_fiscal_year_id')

    fiscal_year = None
    start_date  = None
    end_date    = None

    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()
        if fiscal_year:
            start_date = fiscal_year.start_date
            end_date   = fiscal_year.end_date

    if start_date_str:
        start_date = bs_str_to_ad(start_date_str)
    if end_date_str:
        end_date = bs_str_to_ad(end_date_str)

    def _line_filter(qs):
        if start_date:
            qs = qs.filter(journal_entry__date__gte=start_date)
        if end_date:
            qs = qs.filter(journal_entry__date__lte=end_date)
        return qs

    # Revenue lines — account-level for NFRS disclosure by nature
    revenue_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='REVENUE', is_deleted=False
    ).order_by('code', 'name')
    revenue_lines = []
    total_revenue = Decimal('0.00')
    for acc in revenue_accounts:
        cr = _line_filter(JournalEntryLine.objects.filter(
            account=acc, journal_entry__company=user_company, entry_type='CREDIT'
        )).aggregate(s=Coalesce(Sum('amount'), Decimal('0.00')))['s']
        dr = _line_filter(JournalEntryLine.objects.filter(
            account=acc, journal_entry__company=user_company, entry_type='DEBIT'
        )).aggregate(s=Coalesce(Sum('amount'), Decimal('0.00')))['s']
        net = cr - dr
        if net != Decimal('0.00'):
            revenue_lines.append({'account': acc, 'amount': net})
            total_revenue += net

    # Expense lines — account-level for NFRS disclosure by nature
    expense_accounts = LedgerAccount.objects.filter(
        company=user_company, account_type='EXPENSE', is_deleted=False
    ).order_by('code', 'name')
    expense_lines = []
    total_expenses = Decimal('0.00')
    for acc in expense_accounts:
        dr = _line_filter(JournalEntryLine.objects.filter(
            account=acc, journal_entry__company=user_company, entry_type='DEBIT'
        )).aggregate(s=Coalesce(Sum('amount'), Decimal('0.00')))['s']
        cr = _line_filter(JournalEntryLine.objects.filter(
            account=acc, journal_entry__company=user_company, entry_type='CREDIT'
        )).aggregate(s=Coalesce(Sum('amount'), Decimal('0.00')))['s']
        net = dr - cr
        if net != Decimal('0.00'):
            expense_lines.append({'account': acc, 'amount': net})
            total_expenses += net

    profit_before_tax = total_revenue - total_expenses

    # NFRS 12 (IAS 12): Corporate Income Tax provision
    # Nepal CIT rate: 25% for most companies (15% for special industries)
    cit_rate = Decimal(str(getattr(user_company, 'cit_rate', 25) or 25))
    tax_expense = Decimal('0.00')
    if profit_before_tax > Decimal('0.00'):
        tax_expense = (profit_before_tax * cit_rate / Decimal('100')).quantize(Decimal('0.01'))
    profit_after_tax = profit_before_tax - tax_expense

    context = {
        'revenue_lines':     revenue_lines,
        'expense_lines':     expense_lines,
        'total_revenue':     total_revenue,
        'total_expenses':    total_expenses,
        'profit_before_tax': profit_before_tax,
        'cit_rate':          cit_rate,
        'tax_expense':       tax_expense,
        'net_profit':        profit_after_tax,
        'company':           user_company,
        'fiscal_year':       fiscal_year,
        'start_date':        start_date_str,
        'end_date':          end_date_str,
        'report_title':      'Statement of Profit or Loss',
    }
    return render(request, 'reports/profit_and_loss_report.html', context)


# ═══════════════════════════════════════════════════════════════════════════
# RATIO ANALYSIS REPORT
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def ratio_analysis_report(request, template='reports/ratio_analysis_report.html'):
    """
    Financial ratio analysis covering:
    Liquidity  : Current Ratio, Quick Ratio, Cash Ratio
    Profitability: Gross Margin, Net Margin, ROA, ROE
    Efficiency : AR Turnover, AP Turnover, Inventory Turnover, Asset Turnover
    Leverage   : Debt-to-Equity, Debt Ratio, Interest Coverage (N/A without interest data)
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()

    def account_balance(account_type_list, entry_type):
        """Sum journal entry lines for given account types and entry type."""
        return JournalEntryLine.objects.filter(
            account__company=user_company,
            account__account_type__in=account_type_list,
            entry_type=entry_type,
            journal_entry__is_deleted=False,
            journal_entry__company=user_company,
        ).aggregate(total=Coalesce(Sum('amount'), Decimal('0.00')))['total']

    def net_account_balance(account_type, normal_side):
        """Net balance for an account type (normal_side = 'DEBIT' or 'CREDIT')."""
        debits  = account_balance([account_type], 'DEBIT')
        credits = account_balance([account_type], 'CREDIT')
        if normal_side == 'DEBIT':
            return debits - credits
        return credits - debits

    # ── Balance Sheet items ──────────────────────────────────────────────
    total_assets      = net_account_balance('ASSET',     'DEBIT')
    total_liabilities = net_account_balance('LIABILITY', 'CREDIT')
    total_equity      = net_account_balance('EQUITY',    'CREDIT')

    # Current assets = Cash + Bank + AR
    cash_bank_ids = list(LedgerAccount.objects.filter(
        company=user_company,
        account_type='ASSET',
        name__in=['Cash', 'Bank', 'Accounts Receivable']
    ).values_list('id', flat=True))
    # current_assets computed below via cash_and_bank + ar_balance
    # Simpler approach using outstanding invoices + cash/bank
    ar_balance = Invoice.active_objects.filter(
        company=user_company, outstanding_balance__gt=0
    ).aggregate(t=Coalesce(Sum('outstanding_balance'), Decimal('0')))['t']

    cash_ids = list(LedgerAccount.objects.filter(
        company=user_company, account_type='ASSET', name__in=['Cash', 'Bank']
    ).values_list('id', flat=True))
    cash_balance = JournalEntryLine.objects.filter(
        account_id__in=cash_ids,
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
    ).aggregate(
        debits=Coalesce(Sum('amount', filter=Q(entry_type='DEBIT')),  Decimal('0')),
        credits=Coalesce(Sum('amount', filter=Q(entry_type='CREDIT')), Decimal('0')),
    )
    cash_and_bank = cash_balance['debits'] - cash_balance['credits']
    current_assets_total = cash_and_bank + ar_balance

    # Current liabilities = AP
    ap_ids = list(LedgerAccount.objects.filter(
        company=user_company, account_type='LIABILITY', name='Accounts Payable'
    ).values_list('id', flat=True))
    ap_balance_data = JournalEntryLine.objects.filter(
        account_id__in=ap_ids,
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
    ).aggregate(
        debits=Coalesce(Sum('amount', filter=Q(entry_type='DEBIT')),  Decimal('0')),
        credits=Coalesce(Sum('amount', filter=Q(entry_type='CREDIT')), Decimal('0')),
    )
    current_liabilities = ap_balance_data['credits'] - ap_balance_data['debits']

    # ── P&L items ────────────────────────────────────────────────────────
    revenue_qs = JournalEntryLine.objects.filter(
        account__company=user_company,
        account__account_type='REVENUE',
        entry_type='CREDIT',
        journal_entry__is_deleted=False,
        journal_entry__company=user_company,
    )
    expense_qs = JournalEntryLine.objects.filter(
        account__company=user_company,
        account__account_type='EXPENSE',
        entry_type='DEBIT',
        journal_entry__is_deleted=False,
        journal_entry__company=user_company,
    )
    if fiscal_year:
        revenue_qs = revenue_qs.filter(
            journal_entry__date__gte=fiscal_year.start_date,
            journal_entry__date__lte=fiscal_year.end_date,
        )
        expense_qs = expense_qs.filter(
            journal_entry__date__gte=fiscal_year.start_date,
            journal_entry__date__lte=fiscal_year.end_date,
        )

    total_revenue  = revenue_qs.aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
    total_expenses = expense_qs.aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
    net_income     = total_revenue - total_expenses

    # COGS approximation = Purchase Expense account
    cogs_ids = list(LedgerAccount.objects.filter(
        company=user_company, account_type='EXPENSE',
        name__in=['Purchase Expense', 'Cost of Goods Sold', 'COGS']
    ).values_list('id', flat=True))
    cogs = JournalEntryLine.objects.filter(
        account_id__in=cogs_ids,
        entry_type='DEBIT',
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
    ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
    gross_profit = total_revenue - cogs

    # ── Turnover helpers ─────────────────────────────────────────────────
    total_invoiced = Invoice.active_objects.filter(company=user_company).aggregate(
        t=Coalesce(Sum('total'), Decimal('0'))
    )['t']
    total_purchases = VendorBill.objects.filter(
        vendor__company=user_company
    ).aggregate(t=Coalesce(Sum('total_amount'), Decimal('0')))['t']

    def safe_div(num, den):
        if den and den != 0:
            return round(num / den, 2)
        return None

    def pct(num, den):
        v = safe_div(num, den)
        return round(v * 100, 2) if v is not None else None

    # ── Ratios ───────────────────────────────────────────────────────────
    ratios = {
        # Liquidity
        'current_ratio':    safe_div(current_assets_total, current_liabilities),
        'quick_ratio':      safe_div(cash_and_bank + ar_balance, current_liabilities),
        'cash_ratio':       safe_div(cash_and_bank, current_liabilities),
        # Profitability
        'gross_margin_pct': pct(gross_profit, total_revenue),
        'net_margin_pct':   pct(net_income, total_revenue),
        'roa_pct':          pct(net_income, total_assets),
        'roe_pct':          pct(net_income, total_equity),
        # Efficiency
        'ar_turnover':      safe_div(total_invoiced, ar_balance),
        'ap_turnover':      safe_div(total_purchases, current_liabilities),
        'asset_turnover':   safe_div(total_revenue, total_assets),
        # Leverage
        'debt_to_equity':   safe_div(total_liabilities, total_equity),
        'debt_ratio':       safe_div(total_liabilities, total_assets),
    }

    # Days ratios
    ratios['days_sales_outstanding'] = round(365 / ratios['ar_turnover'], 1) if ratios['ar_turnover'] else None
    ratios['days_payable_outstanding'] = round(365 / ratios['ap_turnover'], 1) if ratios['ap_turnover'] else None

    context = {
        'company': user_company,
        'fiscal_year': fiscal_year,
        'ratios': ratios,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'current_assets': current_assets_total,
        'current_liabilities': current_liabilities,
        'cash_and_bank': cash_and_bank,
        'ar_balance': ar_balance,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'gross_profit': gross_profit,
        'net_income': net_income,
        'cogs': cogs,
    }
    return render(request, template, context)


@login_required
def export_ratio_analysis_excel(request):
    from openpyxl import Workbook
    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    # Re-use the same context logic by calling the view and extracting context
    # Simpler: just redirect to the HTML report for now
    return redirect('reports:ratio_analysis_report')


@login_required
def print_ratio_analysis_report(request):
    """Print-optimised version of ratio analysis."""
    return ratio_analysis_report(request, template='reports/print/ratio_analysis_report_print.html')


# ═══════════════════════════════════════════════════════════════════════════
# AP AGING REPORT
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def ap_aging_report(request):
    """Accounts Payable aging — mirrors AR aging but for vendor bills."""
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    as_of_date_str = request.GET.get('as_of_date')
    as_of_date = bs_str_to_ad(as_of_date_str) if as_of_date_str else date.today()

    # Unpaid vendor bills
    bills = VendorBill.objects.filter(
        vendor__company=user_company,
        status='UNPAID',
    ).select_related('vendor').order_by('due_date')

    aging_data = []
    totals = {'current': Decimal('0'), 'days_1_30': Decimal('0'),
              'days_31_60': Decimal('0'), 'days_61_90': Decimal('0'),
              'days_90_plus': Decimal('0'), 'total': Decimal('0')}

    for bill in bills:
        paid = bill.payments.aggregate(
            t=Coalesce(Sum('amount'), Decimal('0'))
        )['t']
        outstanding = bill.total_amount - paid
        if outstanding <= 0:
            continue

        days_overdue = (as_of_date - bill.due_date).days if bill.due_date else 0

        row = {
            'bill': bill,
            'outstanding': outstanding,
            'days_overdue': days_overdue,
            'current':    outstanding if days_overdue <= 0 else Decimal('0'),
            'days_1_30':  outstanding if 1  <= days_overdue <= 30  else Decimal('0'),
            'days_31_60': outstanding if 31 <= days_overdue <= 60  else Decimal('0'),
            'days_61_90': outstanding if 61 <= days_overdue <= 90  else Decimal('0'),
            'days_90_plus': outstanding if days_overdue > 90 else Decimal('0'),
        }
        aging_data.append(row)
        for k in totals:
            if k != 'total':
                totals[k] += row[k]
        totals['total'] += outstanding

    context = {
        'company': user_company,
        'as_of_date': as_of_date,
        'aging_data': aging_data,
        'totals': totals,
    }
    return render(request, 'reports/ap_aging_report.html', context)


@login_required
def export_ap_aging_excel(request):
    from openpyxl import Workbook
    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')
    return redirect('reports:ap_aging_report')


# ═══════════════════════════════════════════════════════════════════════════
# TAX REPORT (VAT)
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def tax_report(request, template='reports/tax_report.html'):
    """
    VAT / Tax report:
    - Output tax (VAT collected on sales invoices)
    - Input tax (VAT paid on vendor bills — approximated from Purchase Expense)
    - Net tax payable
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()

    start_str = request.GET.get('start_date')
    end_str   = request.GET.get('end_date')
    start_date = bs_str_to_ad(start_str) if start_str else (fiscal_year.start_date if fiscal_year else date(date.today().year, 1, 1))
    end_date   = bs_str_to_ad(end_str)   if end_str   else date.today()

    # Output tax — from Tax Payable account credits
    tax_payable_ids = list(LedgerAccount.objects.filter(
        company=user_company, name='Tax Payable'
    ).values_list('id', flat=True))

    output_tax = JournalEntryLine.objects.filter(
        account_id__in=tax_payable_ids,
        entry_type='CREDIT',
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
        journal_entry__date__gte=start_date,
        journal_entry__date__lte=end_date,
    ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']

    # Input tax — sum stored tax_amount on vendor bills
    purchase_agg = VendorBill.objects.filter(
        vendor__company=user_company,
        bill_date__gte=start_date,
        bill_date__lte=end_date,
    ).aggregate(
        t=Coalesce(Sum('tax_amount'), Decimal('0'))
    )
    input_tax = purchase_agg['t']
    # Legacy bills with no stored tax_amount: fall back to derived estimate
    if input_tax == Decimal('0'):
        total_bills = VendorBill.objects.filter(
            vendor__company=user_company,
            bill_date__gte=start_date,
            bill_date__lte=end_date,
        ).aggregate(t=Coalesce(Sum('total_amount'), Decimal('0')))['t']
        if total_bills > Decimal('0'):
            input_tax = (total_bills * user_company.tax_rate / Decimal('100')).quantize(Decimal('0.01'))

    net_tax_payable = output_tax - input_tax

    # Invoice-level breakdown
    invoices = Invoice.active_objects.filter(
        company=user_company,
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        tax_amount__gt=0,
    ).order_by('transaction_date').values(
        'invoice_number', 'transaction_date', 'subtotal',
        'discount_amount', 'tax_percent', 'tax_amount', 'total'
    )

    context = {
        'company': user_company,
        'fiscal_year': fiscal_year,
        'start_date': start_date,
        'end_date': end_date,
        'output_tax': output_tax,
        'input_tax': input_tax,
        'net_tax_payable': net_tax_payable,
        'invoices': invoices,
        'tax_rate': user_company.tax_rate,
    }
    return render(request, template, context)


@login_required
def export_tax_report_excel(request):
    from openpyxl import Workbook
    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    start_str = request.GET.get('start_date')
    end_str   = request.GET.get('end_date')
    start_date = bs_str_to_ad(start_str) if start_str else date(date.today().year, 1, 1)
    end_date   = bs_str_to_ad(end_str)   if end_str   else date.today()

    invoices = Invoice.active_objects.filter(
        company=user_company,
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        tax_amount__gt=0,
    ).order_by('transaction_date')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tax Report'
    ws.append(['Invoice #', 'Date', 'Subtotal', 'Discount', 'Tax %', 'Tax Amount', 'Total'])
    for inv in invoices:
        ws.append([
            inv.invoice_number,
            str(inv.transaction_date),
            float(inv.subtotal),
            float(inv.discount_amount),
            float(inv.tax_percent),
            float(inv.tax_amount),
            float(inv.total),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tax_report.xlsx"'
    wb.save(response)
    return response


@login_required
def print_tax_report(request):
    return tax_report(request, template='reports/print/tax_report_print.html')


# ═══════════════════════════════════════════════════════════════════════════
# CASH POSITION REPORT
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def cash_position_report(request):
    """
    Cash and bank balances as of a given date, with recent movements.
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    as_of_str  = request.GET.get('as_of_date')
    as_of_date = bs_str_to_ad(as_of_str) if as_of_str else date.today()

    cash_accounts = LedgerAccount.objects.filter(
        company=user_company,
        account_type='ASSET',
        name__in=['Cash', 'Bank'],
        is_deleted=False,
    )

    account_data = []
    total_balance = Decimal('0')

    for acc in cash_accounts:
        lines = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__company=user_company,
            journal_entry__is_deleted=False,
            journal_entry__date__lte=as_of_date,
        ).aggregate(
            debits=Coalesce(Sum('amount', filter=Q(entry_type='DEBIT')),  Decimal('0')),
            credits=Coalesce(Sum('amount', filter=Q(entry_type='CREDIT')), Decimal('0')),
        )
        balance = lines['debits'] - lines['credits']

        # Recent 10 transactions
        recent = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__company=user_company,
            journal_entry__is_deleted=False,
            journal_entry__date__lte=as_of_date,
        ).select_related('journal_entry').order_by('-journal_entry__date')[:10]

        account_data.append({
            'account': acc,
            'balance': balance,
            'recent': recent,
        })
        total_balance += balance

    # Inflows vs outflows for the last 30 days
    thirty_days_ago = date.today() - __import__('datetime').timedelta(days=30)
    cash_ids = list(cash_accounts.values_list('id', flat=True))

    inflows = JournalEntryLine.objects.filter(
        account_id__in=cash_ids,
        entry_type='DEBIT',
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
        journal_entry__date__gte=thirty_days_ago,
        journal_entry__date__lte=as_of_date,
    ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']

    outflows = JournalEntryLine.objects.filter(
        account_id__in=cash_ids,
        entry_type='CREDIT',
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
        journal_entry__date__gte=thirty_days_ago,
        journal_entry__date__lte=as_of_date,
    ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']

    context = {
        'company': user_company,
        'as_of_date': as_of_date,
        'account_data': account_data,
        'total_balance': total_balance,
        'inflows_30d': inflows,
        'outflows_30d': outflows,
        'net_30d': inflows - outflows,
    }
    return render(request, 'reports/cash_position_report.html', context)


@login_required
def export_cash_position_excel(request):
    return redirect('reports:cash_position_report')


# ═══════════════════════════════════════════════════════════════════════════
# PAYROLL SUMMARY REPORT (in reports app)
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def payroll_summary_report(request, template='reports/payroll_summary_report.html'):
    """Payroll summary across all runs for the company."""
    from apps.hrpayroll.models import PayrollRun, Payslip, Employee

    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()

    runs = PayrollRun.active_objects.filter(company=user_company).order_by('-period_start_date')
    if fiscal_year:
        runs = runs.filter(
            period_start_date__gte=fiscal_year.start_date,
            period_end_date__lte=fiscal_year.end_date,
        )

    # Aggregate totals
    totals = runs.aggregate(
        total_gross=Coalesce(Sum('total_gross_pay'), Decimal('0')),
        total_net=Coalesce(Sum('total_net_pay'), Decimal('0')),
    )
    total_deductions = totals['total_gross'] - totals['total_net']

    # Per-employee summary
    employee_summary = Payslip.active_objects.filter(
        payroll_run__company=user_company,
    )
    if fiscal_year:
        employee_summary = employee_summary.filter(
            payroll_run__period_start_date__gte=fiscal_year.start_date,
        )
    employee_summary = employee_summary.values(
        'employee__first_name', 'employee__last_name', 'employee__email'
    ).annotate(
        total_gross=Coalesce(Sum('gross_pay'), Decimal('0')),
        total_deductions=Coalesce(Sum('total_deductions'), Decimal('0')),
        total_net=Coalesce(Sum('net_pay'), Decimal('0')),
        payslip_count=Count('id'),
    ).order_by('employee__last_name', 'employee__first_name')

    active_employees = Employee.active_objects.filter(
        company=user_company, is_active=True
    ).count()

    context = {
        'company': user_company,
        'fiscal_year': fiscal_year,
        'runs': runs,
        'totals': totals,
        'total_deductions': total_deductions,
        'employee_summary': employee_summary,
        'active_employees': active_employees,
        'run_count': runs.count(),
    }
    return render(request, template, context)


@login_required
def export_payroll_summary_excel(request):
    from openpyxl import Workbook
    from apps.hrpayroll.models import PayrollRun

    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    runs = PayrollRun.active_objects.filter(company=user_company).order_by('-period_start_date')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll Summary'
    ws.append(['Period Start', 'Period End', 'Payroll Date', 'Status',
               'Gross Pay', 'Net Pay', 'Deductions'])
    for run in runs:
        ws.append([
            str(run.period_start_date),
            str(run.period_end_date),
            str(run.payroll_date),
            run.status,
            float(run.total_gross_pay),
            float(run.total_net_pay),
            float(run.total_gross_pay - run.total_net_pay),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payroll_summary.xlsx"'
    wb.save(response)
    return response


@login_required
def print_payroll_summary_report(request):
    return payroll_summary_report(request, template='reports/print/payroll_summary_report_print.html')



# ═══════════════════════════════════════════════════════════════════════════
# VAT PERIOD REPORT  (monthly / 4-monthly / bi-yearly)
# ═══════════════════════════════════════════════════════════════════════════

def _vat_period_label(period_type, period_index, fiscal_year):
    """
    Return a human-readable label for a VAT filing period.

    period_type : 'monthly' | '4monthly' | 'biyearly'
    period_index: 0-based index within the fiscal year
    fiscal_year : FiscalYear instance (used for year label)
    """
    fy_label = fiscal_year.name if fiscal_year else ''
    if period_type == 'monthly':
        month_names = [
            'Month 1', 'Month 2', 'Month 3', 'Month 4',
            'Month 5', 'Month 6', 'Month 7', 'Month 8',
            'Month 9', 'Month 10', 'Month 11', 'Month 12',
        ]
        label = month_names[period_index] if period_index < len(month_names) else f'Month {period_index + 1}'
    elif period_type == '4monthly':
        labels = ['Period 1 (Months 1–4)', 'Period 2 (Months 5–8)', 'Period 3 (Months 9–12)']
        label = labels[period_index] if period_index < len(labels) else f'Period {period_index + 1}'
    else:  # biyearly
        labels = ['Half 1 (Months 1–6)', 'Half 2 (Months 7–12)']
        label = labels[period_index] if period_index < len(labels) else f'Half {period_index + 1}'
    return f'{label} — {fy_label}' if fy_label else label


def _build_vat_periods(start_date, end_date, period_type):
    """
    Slice the date range [start_date, end_date] into sub-periods.

    Returns a list of (period_start, period_end) date tuples.
    For 'monthly'  → one entry per calendar month within the range.
    For '4monthly' → groups of ~4 months (122 days each, last absorbs remainder).
    For 'biyearly' → two halves split at the midpoint.
    """
    from datetime import timedelta
    import calendar

    periods = []

    if period_type == 'monthly':
        current = start_date.replace(day=1)
        while current <= end_date:
            last_day = calendar.monthrange(current.year, current.month)[1]
            p_start = max(current, start_date)
            p_end   = min(date(current.year, current.month, last_day), end_date)
            if p_start <= p_end:
                periods.append((p_start, p_end))
            # advance to next month
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)

    elif period_type == '4monthly':
        total_days = (end_date - start_date).days + 1
        chunk = total_days // 3
        for i in range(3):
            p_start = start_date + timedelta(days=i * chunk)
            if i == 2:
                p_end = end_date
            else:
                p_end = start_date + timedelta(days=(i + 1) * chunk - 1)
            if p_start <= end_date:
                periods.append((p_start, min(p_end, end_date)))

    else:  # biyearly
        total_days = (end_date - start_date).days + 1
        mid = start_date + timedelta(days=total_days // 2 - 1)
        periods.append((start_date, mid))
        periods.append((mid + timedelta(days=1), end_date))

    return periods


def _compute_vat_for_period(user_company, p_start, p_end, tax_payable_ids):
    """
    Compute output VAT (from sales invoices) and input VAT (from vendor bills)
    for a single date period.  Returns a dict with all figures needed for display.
    """
    # ── Output VAT: sum of tax_amount on invoices in this period ──────────
    sales_qs = Invoice.active_objects.filter(
        company=user_company,
        transaction_date__gte=p_start,
        transaction_date__lte=p_end,
        tax_amount__gt=0,
    ).aggregate(
        taxable_sales=Coalesce(Sum('subtotal'),      Decimal('0')),
        discount_total=Coalesce(Sum('discount_amount'), Decimal('0')),
        output_vat=Coalesce(Sum('tax_amount'),       Decimal('0')),
        gross_sales=Coalesce(Sum('total'),           Decimal('0')),
    )

    # Also pull from Tax Payable ledger credits as a cross-check figure
    ledger_output_vat = JournalEntryLine.objects.filter(
        account_id__in=tax_payable_ids,
        entry_type='CREDIT',
        journal_entry__company=user_company,
        journal_entry__is_deleted=False,
        journal_entry__date__gte=p_start,
        journal_entry__date__lte=p_end,
    ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']

    # ── Input VAT: vendor bills in this period ────────────────────────────
    purchase_qs = VendorBill.objects.filter(
        vendor__company=user_company,
        bill_date__gte=p_start,
        bill_date__lte=p_end,
    ).aggregate(
        total_purchases=Coalesce(Sum('total_amount'), Decimal('0')),
        input_vat=Coalesce(Sum('tax_amount'), Decimal('0')),
    )
    total_purchases = purchase_qs['total_purchases']
    input_vat = purchase_qs['input_vat']
    taxable_purchases = (total_purchases - input_vat).quantize(Decimal('0.01'))

    # Legacy bills (before tax_amount field) have tax_amount=0 but may have
    # tax embedded in total. For those rows, fall back to derivation only when
    # the entire period has input_vat=0 yet purchases exist.
    if input_vat == Decimal('0') and total_purchases > Decimal('0'):
        tax_rate = user_company.tax_rate
        if getattr(user_company, 'vat_inclusive', False):
            input_vat = (total_purchases * tax_rate / (Decimal('100') + tax_rate)).quantize(Decimal('0.01'))
        else:
            input_vat = (total_purchases * tax_rate / Decimal('100')).quantize(Decimal('0.01'))
        taxable_purchases = total_purchases - input_vat

    output_vat = sales_qs['output_vat']
    net_vat    = output_vat - input_vat

    return {
        'period_start':      p_start,
        'period_end':        p_end,
        'taxable_sales':     sales_qs['taxable_sales'],
        'discount_total':    sales_qs['discount_total'],
        'gross_sales':       sales_qs['gross_sales'],
        'output_vat':        output_vat,
        'ledger_output_vat': ledger_output_vat,
        'taxable_purchases': taxable_purchases,
        'total_purchases':   total_purchases,
        'input_vat':         input_vat,
        'net_vat':           net_vat,
        'is_payable':        net_vat >= Decimal('0'),
    }


@login_required
def vat_period_report(request, template='reports/vat_period_report.html'):
    """
    VAT Period Report.

    Breaks the selected date range into filing periods (monthly, 4-monthly,
    or bi-yearly) and shows for each period:
      - Output VAT  : VAT collected on sales invoices
      - Input VAT   : VAT paid on vendor bills (derived from bill totals)
      - Net VAT     : Output − Input  (positive = payable, negative = refund)

    Query params:
      start_date   BS date string  (defaults to FY start or Jan 1)
      end_date     BS date string  (defaults to today)
      period_type  monthly | 4monthly | biyearly  (default: monthly)
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(
            id=fiscal_year_id, company=user_company
        ).first()

    # ── Date range ────────────────────────────────────────────────────────
    start_str  = request.GET.get('start_date')
    end_str    = request.GET.get('end_date')
    period_type = request.GET.get('period_type', 'monthly')
    if period_type not in ('monthly', '4monthly', 'biyearly'):
        period_type = 'monthly'

    default_start = fiscal_year.start_date if fiscal_year else date(date.today().year, 1, 1)
    default_end   = fiscal_year.end_date   if fiscal_year else date.today()

    start_date = bs_str_to_ad(start_str) if start_str else default_start
    end_date   = bs_str_to_ad(end_str)   if end_str   else default_end

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # ── Tax Payable ledger account IDs (for cross-check) ─────────────────
    tax_payable_ids = list(
        LedgerAccount.objects.filter(
            company=user_company, name='Tax Payable'
        ).values_list('id', flat=True)
    )

    # ── Build periods and compute VAT for each ────────────────────────────
    raw_periods = _build_vat_periods(start_date, end_date, period_type)
    periods = []
    for idx, (p_start, p_end) in enumerate(raw_periods):
        data = _compute_vat_for_period(user_company, p_start, p_end, tax_payable_ids)
        data['label'] = _vat_period_label(period_type, idx, fiscal_year)
        periods.append(data)

    # ── Grand totals ──────────────────────────────────────────────────────
    totals = {
        'taxable_sales':     sum(p['taxable_sales']     for p in periods),
        'discount_total':    sum(p['discount_total']    for p in periods),
        'gross_sales':       sum(p['gross_sales']       for p in periods),
        'output_vat':        sum(p['output_vat']        for p in periods),
        'taxable_purchases': sum(p['taxable_purchases'] for p in periods),
        'total_purchases':   sum(p['total_purchases']   for p in periods),
        'input_vat':         sum(p['input_vat']         for p in periods),
        'net_vat':           sum(p['net_vat']           for p in periods),
    }

    # ── Invoice-level detail for the full range ───────────────────────────
    invoice_detail = Invoice.active_objects.filter(
        company=user_company,
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        tax_amount__gt=0,
    ).order_by('transaction_date').values(
        'invoice_number', 'transaction_date', 'subtotal',
        'discount_amount', 'tax_percent', 'tax_amount', 'total',
    )

    # ── Vendor bill detail for the full range ─────────────────────────────
    from django.db.models import ExpressionWrapper, F as _F, DecimalField as _DF
    bill_detail = VendorBill.objects.filter(
        vendor__company=user_company,
        bill_date__gte=start_date,
        bill_date__lte=end_date,
    ).select_related('vendor').annotate(
        subtotal_amount=ExpressionWrapper(
            _F('total_amount') - _F('tax_amount'), output_field=_DF(max_digits=10, decimal_places=2)
        )
    ).order_by('bill_date').values(
        'bill_number', 'bill_date', 'total_amount', 'tax_amount', 'tax_percent',
        'subtotal_amount', 'vendor__name',
    )

    logger.info(
        'VAT_PERIOD_REPORT actor=%s company=%s period_type=%s start=%s end=%s',
        request.user.email, user_company, period_type, start_date, end_date,
    )

    context = {
        'company':      user_company,
        'fiscal_year':  fiscal_year,
        'start_date':   start_date,
        'end_date':     end_date,
        'period_type':  period_type,
        'periods':      periods,
        'totals':       totals,
        'invoice_detail': invoice_detail,
        'bill_detail':    bill_detail,
        'tax_rate':     user_company.tax_rate,
    }
    return render(request, template, context)


@login_required
def export_vat_period_report_excel(request):
    """Export the VAT period report as an Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(
            id=fiscal_year_id, company=user_company
        ).first()

    start_str   = request.GET.get('start_date')
    end_str     = request.GET.get('end_date')
    period_type = request.GET.get('period_type', 'monthly')
    if period_type not in ('monthly', '4monthly', 'biyearly'):
        period_type = 'monthly'

    default_start = fiscal_year.start_date if fiscal_year else date(date.today().year, 1, 1)
    default_end   = fiscal_year.end_date   if fiscal_year else date.today()

    start_date = bs_str_to_ad(start_str) if start_str else default_start
    end_date   = bs_str_to_ad(end_str)   if end_str   else default_end
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    tax_payable_ids = list(
        LedgerAccount.objects.filter(
            company=user_company, name='Tax Payable'
        ).values_list('id', flat=True)
    )

    raw_periods = _build_vat_periods(start_date, end_date, period_type)
    periods = []
    for idx, (p_start, p_end) in enumerate(raw_periods):
        data = _compute_vat_for_period(user_company, p_start, p_end, tax_payable_ids)
        data['label'] = _vat_period_label(period_type, idx, fiscal_year)
        periods.append(data)

    wb = Workbook()

    # ── Sheet 1: Period Summary ───────────────────────────────────────────
    ws = wb.active
    ws.title = 'VAT Period Summary'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='166534')   # green-800
    total_fill   = PatternFill('solid', fgColor='DCFCE7')   # green-100
    center_align = Alignment(horizontal='center')

    ws.append([
        f'VAT Period Report — {user_company.name}',
    ])
    ws.append([f'Period: {start_date} to {end_date}  |  Filing: {period_type}'])
    ws.append([f'Tax Rate: {user_company.tax_rate}%'])
    ws.append([])

    col_headers = [
        'Period', 'From', 'To',
        'Taxable Sales', 'Discount', 'Gross Sales', 'Output VAT',
        'Taxable Purchases', 'Total Purchases', 'Input VAT',
        'Net VAT Payable',
    ]
    ws.append(col_headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(col_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center_align

    for p in periods:
        ws.append([
            p['label'],
            str(p['period_start']),
            str(p['period_end']),
            float(p['taxable_sales']),
            float(p['discount_total']),
            float(p['gross_sales']),
            float(p['output_vat']),
            float(p['taxable_purchases']),
            float(p['total_purchases']),
            float(p['input_vat']),
            float(p['net_vat']),
        ])

    # Totals row
    totals_row_data = [
        'TOTAL', '', '',
        float(sum(p['taxable_sales']     for p in periods)),
        float(sum(p['discount_total']    for p in periods)),
        float(sum(p['gross_sales']       for p in periods)),
        float(sum(p['output_vat']        for p in periods)),
        float(sum(p['taxable_purchases'] for p in periods)),
        float(sum(p['total_purchases']   for p in periods)),
        float(sum(p['input_vat']         for p in periods)),
        float(sum(p['net_vat']           for p in periods)),
    ]
    ws.append(totals_row_data)
    total_row_idx = ws.max_row
    for col_idx in range(1, len(col_headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = total_fill

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # ── Sheet 2: Invoice Detail ───────────────────────────────────────────
    ws2 = wb.create_sheet('Sales Invoice Detail')
    ws2.append(['Invoice #', 'Date', 'Subtotal', 'Discount', 'Tax %', 'Tax Amount', 'Total'])
    inv_header_row = ws2.max_row
    for col_idx in range(1, 8):
        cell = ws2.cell(row=inv_header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    invoices = Invoice.active_objects.filter(
        company=user_company,
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        tax_amount__gt=0,
    ).order_by('transaction_date')

    for inv in invoices:
        ws2.append([
            inv.invoice_number,
            str(inv.transaction_date),
            float(inv.subtotal),
            float(inv.discount_amount),
            float(inv.tax_percent),
            float(inv.tax_amount),
            float(inv.total),
        ])

    # ── Sheet 3: Vendor Bill Detail ───────────────────────────────────────
    ws3 = wb.create_sheet('Purchase Bill Detail')
    ws3.append(['Bill #', 'Date', 'Vendor', 'Subtotal (excl. VAT)', 'VAT %', 'Input VAT', 'Bill Total'])
    bill_header_row = ws3.max_row
    for col_idx in range(1, 8):
        cell = ws3.cell(row=bill_header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    bills = VendorBill.objects.filter(
        vendor__company=user_company,
        bill_date__gte=start_date,
        bill_date__lte=end_date,
    ).select_related('vendor').order_by('bill_date')

    for bill in bills:
        subtotal_ex_vat = float((bill.total_amount - bill.tax_amount).quantize(Decimal('0.01')))
        ws3.append([
            bill.bill_number,
            str(bill.bill_date),
            bill.vendor.name,
            subtotal_ex_vat,
            float(bill.tax_percent),
            float(bill.tax_amount),
            float(bill.total_amount),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vat_period_report.xlsx"'
    wb.save(response)
    return response


@login_required
def print_vat_period_report(request):
    return vat_period_report(request, template='reports/print/vat_period_report_print.html')


# ═══════════════════════════════════════════════════════════════════════════
# NFRS 1 — Statement of Changes in Equity
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def changes_in_equity_report(request):
    """
    NFRS 1 (IAS 1) para 106 — Statement of Changes in Equity.

    Presents movements in each equity component for the period:
      Share Capital, Retained Earnings, Other Equity accounts.

    Derivation logic (no new model needed):
      Opening balance = LedgerOpeningBalance for the active FY
      + movements via JournalEntryLines during the FY
      = Closing balance
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    fiscal_year_id = request.session.get('active_fiscal_year_id')
    fiscal_year = None
    if fiscal_year_id:
        fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id, company=user_company).first()
    if not fiscal_year:
        fiscal_year = FiscalYear.objects.filter(
            company=user_company,
            start_date__lte=date.today(),
            end_date__gte=date.today(),
        ).first()

    equity_accounts = LedgerAccount.objects.filter(
        company=user_company,
        account_type='EQUITY',
        is_deleted=False,
    ).order_by('code', 'name')

    rows = []
    total_opening = Decimal('0.00')
    total_additions = Decimal('0.00')
    total_reductions = Decimal('0.00')
    total_closing = Decimal('0.00')

    for acc in equity_accounts:
        # Opening balance from LedgerOpeningBalance
        ob = Decimal('0.00')
        if fiscal_year:
            lob = LedgerOpeningBalance.objects.filter(
                account=acc, fiscal_year=fiscal_year
            ).first()
            if lob:
                ob = lob.amount if lob.opening_type == 'CREDIT' else -lob.amount

        # Period movements
        qs_base = JournalEntryLine.objects.filter(
            account=acc,
            journal_entry__company=user_company,
            journal_entry__is_deleted=False,
        )
        if fiscal_year:
            qs_base = qs_base.filter(
                journal_entry__date__gte=fiscal_year.start_date,
                journal_entry__date__lte=fiscal_year.end_date,
            )

        agg = qs_base.aggregate(
            cr=Coalesce(Sum('amount', filter=Q(entry_type='CREDIT')), Decimal('0')),
            dr=Coalesce(Sum('amount', filter=Q(entry_type='DEBIT')),  Decimal('0')),
        )
        additions  = agg['cr']   # credits increase equity
        reductions = agg['dr']   # debits decrease equity
        closing = ob + additions - reductions

        rows.append({
            'account':    acc,
            'opening':    ob,
            'additions':  additions,
            'reductions': reductions,
            'closing':    closing,
        })
        total_opening    += ob
        total_additions  += additions
        total_reductions += reductions
        total_closing    += closing

    # Add current period net income as a movement in Retained Earnings
    if fiscal_year:
        rev = JournalEntryLine.objects.filter(
            account__company=user_company,
            account__account_type='REVENUE',
            entry_type='CREDIT',
            journal_entry__is_deleted=False,
            journal_entry__date__gte=fiscal_year.start_date,
            journal_entry__date__lte=fiscal_year.end_date,
        ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
        exp = JournalEntryLine.objects.filter(
            account__company=user_company,
            account__account_type='EXPENSE',
            entry_type='DEBIT',
            journal_entry__is_deleted=False,
            journal_entry__date__gte=fiscal_year.start_date,
            journal_entry__date__lte=fiscal_year.end_date,
        ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
        net_income = rev - exp
    else:
        net_income = Decimal('0.00')

    context = {
        'company':           user_company,
        'fiscal_year':       fiscal_year,
        'rows':              rows,
        'net_income':        net_income,
        'total_opening':     total_opening,
        'total_additions':   total_additions,
        'total_reductions':  total_reductions,
        'total_closing':     total_closing + net_income,
        'report_title':      'Statement of Changes in Equity',
    }
    return render(request, 'reports/changes_in_equity_report.html', context)


# ═══════════════════════════════════════════════════════════════════════════
# NFRS 13 — Fixed Asset Register Report
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def fixed_asset_register_report(request):
    """NFRS 13 (IAS 16) fixed asset register with NBV and depreciation schedule."""
    from apps.bookkeeping.models import FixedAsset
    from apps.bookkeeping.fixed_asset_service import depreciation_schedule

    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    assets = FixedAsset.objects.filter(
        company=user_company,
        is_deleted=False,
    ).select_related('asset_account').order_by('category', 'name')

    asset_rows = []
    total_cost = Decimal('0.00')
    total_accum_dep = Decimal('0.00')
    total_nbv = Decimal('0.00')

    for asset in assets:
        asset_rows.append({
            'asset':                  asset,
            'schedule':               depreciation_schedule(asset),
        })
        total_cost      += asset.cost
        total_accum_dep += asset.accumulated_depreciation
        total_nbv       += asset.net_book_value

    context = {
        'company':          user_company,
        'asset_rows':       asset_rows,
        'total_cost':       total_cost,
        'total_accum_dep':  total_accum_dep,
        'total_nbv':        total_nbv,
        'report_title':     'Fixed Asset Register',
    }
    return render(request, 'reports/fixed_asset_register_report.html', context)


# ── Company Admin: Report Access Management ───────────────────────────────────

@login_required
def report_access_manage(request):
    """Company admin view to assign reports to users."""
    user = request.user
    company = request.user_company
    if not (user.is_superuser or user.is_company_admin):
        messages.error(request, "Only company admins can manage report access.")
        return redirect('reports:report_dashboard')

    from apps.accounts.models import User as AppUser
    from apps.reports.report_registry import get_enabled_reports, REPORT_REGISTRY, SECTION_ORDER

    company_users = AppUser.objects.filter(
        company=company, is_active=True, is_deleted=False
    ).exclude(pk=user.pk).exclude(is_superuser=True).order_by('email')

    enabled_slugs = get_enabled_reports(company)
    enabled_reports = [
        (slug, REPORT_REGISTRY[slug])
        for slug in REPORT_REGISTRY
        if slug in enabled_slugs
    ]

    all_access = UserReportAccess.objects.filter(
        company=company, is_deleted=False
    ).values_list('user_id', 'report_name')
    raw_map = {}
    for uid, rname in all_access:
        raw_map.setdefault(uid, set()).add(rname)

    # Annotate each user with their granted set so the template can access it directly
    users_with_access = []
    for u in company_users:
        u.granted_reports = raw_map.get(u.pk, set())
        users_with_access.append(u)

    return render(request, 'reports/report_access.html', {
        'company_users': users_with_access,
        'enabled_reports': enabled_reports,
    })


@login_required
def report_access_toggle(request, user_pk):
    """Toggle a single report on/off for a user. POST only."""
    from django.views.decorators.http import require_POST
    if request.method != 'POST':
        return redirect('reports:report_access_manage')

    admin_user = request.user
    company = request.user_company
    if not (admin_user.is_superuser or admin_user.is_company_admin):
        messages.error(request, "Permission denied.")
        return redirect('reports:report_dashboard')

    from apps.accounts.models import User as AppUser
    target_user = get_object_or_404(AppUser, pk=user_pk, company=company)
    report_name = request.POST.get('report_name', '').strip()

    if report_name not in REPORT_REGISTRY:
        messages.error(request, f"Unknown report: {report_name}")
        return redirect('reports:report_access_manage')

    obj, created = UserReportAccess.objects.get_or_create(
        company=company, user=target_user, report_name=report_name,
        defaults={'granted_by': admin_user, 'created_by': admin_user},
    )
    if not created:
        # already existed — soft-delete toggles it off, or restore if deleted
        if obj.is_deleted:
            obj.is_deleted = False
            obj.granted_by = admin_user
            obj.save(update_fields=['is_deleted', 'granted_by'])
        else:
            obj.is_deleted = True
            obj.save(update_fields=['is_deleted'])

    return redirect('reports:report_access_manage')


@login_required
def report_access_bulk(request, user_pk):
    """Grant or revoke ALL enabled reports for a user at once. POST only."""
    if request.method != 'POST':
        return redirect('reports:report_access_manage')

    admin_user = request.user
    company = request.user_company
    if not (admin_user.is_superuser or admin_user.is_company_admin):
        messages.error(request, "Permission denied.")
        return redirect('reports:report_dashboard')

    from apps.accounts.models import User as AppUser
    from apps.reports.report_registry import get_enabled_reports
    target_user = get_object_or_404(AppUser, pk=user_pk, company=company)
    action = request.POST.get('action')  # 'grant_all' or 'revoke_all'
    enabled = get_enabled_reports(company)

    if action == 'grant_all':
        for slug in enabled:
            UserReportAccess.objects.update_or_create(
                company=company, user=target_user, report_name=slug,
                defaults={'granted_by': admin_user, 'is_deleted': False, 'created_by': admin_user},
            )
        messages.success(request, f"Granted all reports to {target_user.email}.")
    elif action == 'revoke_all':
        UserReportAccess.objects.filter(
            company=company, user=target_user
        ).update(is_deleted=True)
        messages.success(request, f"Revoked all reports from {target_user.email}.")

    return redirect('reports:report_access_manage')


# ═══════════════════════════════════════════════════════════════════════════
# VENDOR-WISE INVENTORY REPORT
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def vendor_wise_inventory_report(request):
    """
    Inventory grouped by preferred vendor.
    Shows each vendor's products with current stock, cost price, and total stock value.
    Products with no vendor assigned are collected under a 'No Vendor' group.
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    products_qs = (
        Product.active_objects
        .filter(company=user_company, is_service=False)
        .select_related('vendor', 'category', 'productstock')
        .order_by('vendor__name', 'name')
    )

    # Filter by vendor if requested
    selected_vendor_id = request.GET.get('vendor', '').strip()
    if selected_vendor_id:
        products_qs = products_qs.filter(vendor_id=selected_vendor_id)

    # Group products by vendor in Python (avoids complex ORM grouping)
    vendor_groups = {}  # vendor_id (or None) -> {'vendor': obj|None, 'products': [], totals}
    for product in products_qs:
        try:
            stock_obj = product.productstock
            current_stock = stock_obj.stock
            minimum_stock = stock_obj.minimum_stock
        except Exception:
            current_stock = 0
            minimum_stock = 0

        stock_value = (Decimal(str(current_stock)) * product.cost_price).quantize(Decimal('0.01'))
        low_stock = current_stock <= minimum_stock

        vendor = product.vendor
        key = str(vendor.id) if vendor else '__none__'
        if key not in vendor_groups:
            vendor_groups[key] = {
                'vendor': vendor,
                'products': [],
                'total_stock_value': Decimal('0.00'),
                'total_products': 0,
                'low_stock_count': 0,
            }
        vendor_groups[key]['products'].append({
            'product': product,
            'current_stock': current_stock,
            'minimum_stock': minimum_stock,
            'stock_value': stock_value,
            'low_stock': low_stock,
        })
        vendor_groups[key]['total_stock_value'] += stock_value
        vendor_groups[key]['total_products'] += 1
        if low_stock:
            vendor_groups[key]['low_stock_count'] += 1

    # Sort: named vendors alphabetically, 'No Vendor' last
    sorted_groups = sorted(
        vendor_groups.values(),
        key=lambda g: (g['vendor'] is None, g['vendor'].name.lower() if g['vendor'] else '')
    )

    vendors = Vendor.active_objects.filter(company=user_company).order_by('name')
    grand_total_value = sum(g['total_stock_value'] for g in sorted_groups)

    context = {
        'vendor_groups': sorted_groups,
        'vendors': vendors,
        'selected_vendor_id': selected_vendor_id,
        'grand_total_value': grand_total_value,
        'company': user_company,
    }
    return render(request, 'reports/products/vendor_wise_inventory_report.html', context)


@login_required
def export_vendor_wise_inventory_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    products_qs = (
        Product.active_objects
        .filter(company=user_company, is_service=False)
        .select_related('vendor', 'category', 'productstock')
        .order_by('vendor__name', 'name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendor-wise Inventory'

    headers = ['Vendor', 'Product', 'SKU', 'Category', 'Cost Price', 'Selling Price', 'Stock', 'Min Stock', 'Stock Value', 'Low Stock']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for product in products_qs:
        try:
            ps = product.productstock
            stock = ps.stock
            min_stock = ps.minimum_stock
        except Exception:
            stock = 0
            min_stock = 0
        stock_value = float(Decimal(str(stock)) * product.cost_price)
        ws.append([
            product.vendor.name if product.vendor else 'No Vendor',
            product.name,
            product.sku or '',
            product.category.name if product.category else '',
            float(product.cost_price),
            float(product.price),
            stock,
            min_stock,
            stock_value,
            'Yes' if stock <= min_stock else 'No',
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vendor_wise_inventory.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════
# SALES BY REFERENCE NUMBER REPORT
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def sales_by_reference_report(request):
    """
    Sales invoices grouped by reference_number.
    Useful for tracking sales against a customer PO#, contract#, or cheque#.
    Invoices with no reference are shown under 'No Reference'.
    """
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_str = request.GET.get('start_date', '').strip()
    end_str   = request.GET.get('end_date', '').strip()
    status_filter = request.GET.get('status', '').strip()
    ref_search = request.GET.get('reference', '').strip()

    invoices_qs = Invoice.active_objects.filter(
        company=user_company
    ).exclude(status='CANCELLED').select_related('customer').order_by('reference_number', 'transaction_date')

    if start_str:
        start_date = bs_str_to_ad(start_str)
        if start_date:
            invoices_qs = invoices_qs.filter(transaction_date__gte=start_date)
    if end_str:
        end_date = bs_str_to_ad(end_str)
        if end_date:
            invoices_qs = invoices_qs.filter(transaction_date__lte=end_date)
    if status_filter:
        invoices_qs = invoices_qs.filter(status=status_filter)
    if ref_search:
        invoices_qs = invoices_qs.filter(reference_number__icontains=ref_search)

    # Group by reference_number in Python
    ref_groups = {}
    for inv in invoices_qs:
        ref = inv.reference_number or ''
        key = ref if ref else '__none__'
        if key not in ref_groups:
            ref_groups[key] = {
                'reference': ref if ref else None,
                'invoices': [],
                'total_amount': Decimal('0.00'),
                'invoice_count': 0,
            }
        ref_groups[key]['invoices'].append(inv)
        ref_groups[key]['total_amount'] += inv.total
        ref_groups[key]['invoice_count'] += 1

    # Sort: named references alphabetically, 'No Reference' last
    sorted_groups = sorted(
        ref_groups.values(),
        key=lambda g: (g['reference'] is None, (g['reference'] or '').lower())
    )

    grand_total = sum(g['total_amount'] for g in sorted_groups)
    grand_count = sum(g['invoice_count'] for g in sorted_groups)

    context = {
        'ref_groups': sorted_groups,
        'grand_total': grand_total,
        'grand_count': grand_count,
        'start_date': start_str,
        'end_date': end_str,
        'status_filter': status_filter,
        'ref_search': ref_search,
        'company': user_company,
    }
    return render(request, 'reports/sales/sales_by_reference_report.html', context)


@login_required
def export_sales_by_reference_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    start_str = request.GET.get('start_date', '').strip()
    end_str   = request.GET.get('end_date', '').strip()

    invoices_qs = Invoice.active_objects.filter(
        company=user_company
    ).exclude(status='CANCELLED').select_related('customer').order_by('reference_number', 'transaction_date')

    if start_str:
        sd = bs_str_to_ad(start_str)
        if sd:
            invoices_qs = invoices_qs.filter(transaction_date__gte=sd)
    if end_str:
        ed = bs_str_to_ad(end_str)
        if ed:
            invoices_qs = invoices_qs.filter(transaction_date__lte=ed)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales by Reference'

    headers = ['Reference No.', 'Invoice #', 'Date', 'Customer', 'Status', 'Subtotal', 'Discount', 'Tax', 'Total']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for inv in invoices_qs:
        ws.append([
            inv.reference_number or 'No Reference',
            inv.invoice_number or '',
            str(inv.transaction_date),
            inv.customer.name if inv.customer else 'Cash Sale',
            inv.status,
            float(inv.subtotal),
            float(inv.discount_amount),
            float(inv.tax_amount),
            float(inv.total),
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_by_reference.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════
# PURCHASE PRICE HISTORY REPORT
# (same product purchased on different dates at different cost prices)
# ═══════════════════════════════════════════════════════════════════════════

@login_required
def purchase_price_history_report(request):
    """
    Shows all VendorBillItem records per product so you can see how cost
    price has changed across purchases on different dates from different vendors.
    Includes the current cost_price on the product for comparison.
    """
    from apps.billing.models import VendorBillItem

    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company. Please contact an administrator.")
        return redirect('accounts:user_dashboard')

    start_str = request.GET.get('start_date', '').strip()
    end_str   = request.GET.get('end_date', '').strip()
    selected_product_id = request.GET.get('product', '').strip()
    selected_vendor_id  = request.GET.get('vendor', '').strip()

    bill_items_qs = (
        VendorBillItem.active_objects
        .filter(
            product__company=user_company,
            product__isnull=False,
        )
        .select_related('vendor_bill', 'vendor_bill__vendor', 'product', 'product__category')
        .order_by('product__name', 'vendor_bill__bill_date')
    )

    if start_str:
        sd = bs_str_to_ad(start_str)
        if sd:
            bill_items_qs = bill_items_qs.filter(vendor_bill__bill_date__gte=sd)
    if end_str:
        ed = bs_str_to_ad(end_str)
        if ed:
            bill_items_qs = bill_items_qs.filter(vendor_bill__bill_date__lte=ed)
    if selected_product_id:
        bill_items_qs = bill_items_qs.filter(product_id=selected_product_id)
    if selected_vendor_id:
        bill_items_qs = bill_items_qs.filter(vendor_bill__vendor_id=selected_vendor_id)

    # Group by product
    product_groups = {}
    for item in bill_items_qs:
        pid = str(item.product_id)
        if pid not in product_groups:
            product_groups[pid] = {
                'product': item.product,
                'purchases': [],
                'min_price': item.price,
                'max_price': item.price,
                'total_qty': 0,
                'total_cost': Decimal('0.00'),
            }
        grp = product_groups[pid]
        grp['purchases'].append(item)
        grp['min_price'] = min(grp['min_price'], item.price)
        grp['max_price'] = max(grp['max_price'], item.price)
        grp['total_qty'] += item.quantity
        grp['total_cost'] += item.quantity * item.price

    # Weighted average per product group
    for grp in product_groups.values():
        if grp['total_qty']:
            grp['weighted_avg_cost'] = (grp['total_cost'] / Decimal(str(grp['total_qty']))).quantize(Decimal('0.01'))
        else:
            grp['weighted_avg_cost'] = Decimal('0.00')

    sorted_groups = sorted(product_groups.values(), key=lambda g: g['product'].name.lower())

    products = Product.active_objects.filter(company=user_company, is_service=False).order_by('name')
    vendors  = Vendor.active_objects.filter(company=user_company).order_by('name')

    context = {
        'product_groups': sorted_groups,
        'products': products,
        'vendors': vendors,
        'selected_product_id': selected_product_id,
        'selected_vendor_id': selected_vendor_id,
        'start_date': start_str,
        'end_date': end_str,
        'company': user_company,
    }
    return render(request, 'reports/products/purchase_price_history_report.html', context)


@login_required
def export_purchase_price_history_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from apps.billing.models import VendorBillItem

    user_company = request.user_company
    if not user_company:
        return redirect('accounts:user_dashboard')

    start_str = request.GET.get('start_date', '').strip()
    end_str   = request.GET.get('end_date', '').strip()

    bill_items_qs = (
        VendorBillItem.active_objects
        .filter(product__company=user_company, product__isnull=False)
        .select_related('vendor_bill', 'vendor_bill__vendor', 'product')
        .order_by('product__name', 'vendor_bill__bill_date')
    )

    if start_str:
        sd = bs_str_to_ad(start_str)
        if sd:
            bill_items_qs = bill_items_qs.filter(vendor_bill__bill_date__gte=sd)
    if end_str:
        ed = bs_str_to_ad(end_str)
        if ed:
            bill_items_qs = bill_items_qs.filter(vendor_bill__bill_date__lte=ed)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Purchase Price History'

    headers = ['Product', 'SKU', 'Current Cost Price', 'Bill Date', 'Vendor', 'Bill #', 'Qty', 'Unit Cost', 'Line Total']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in bill_items_qs:
        ws.append([
            item.product.name,
            item.product.sku or '',
            float(item.product.cost_price),
            str(item.vendor_bill.bill_date),
            item.vendor_bill.vendor.name,
            item.vendor_bill.bill_number,
            item.quantity,
            float(item.price),
            float(item.quantity * item.price),
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchase_price_history.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# DAYBOOK REPORT
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def daybook_report(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    from apps.utils.nepali_date import today_bs, bs_str_to_ad, ad_date_to_bs_str
    import datetime

    # Date selection — default to today
    date_bs_str = request.GET.get('date_bs', '').strip()
    date_ad_str = request.GET.get('date_ad', '').strip()

    selected_date = None
    selected_date_bs = ''

    if date_bs_str:
        try:
            selected_date = bs_str_to_ad(date_bs_str)
            selected_date_bs = date_bs_str
        except Exception:
            pass
    elif date_ad_str:
        try:
            selected_date = datetime.date.fromisoformat(date_ad_str)
            selected_date_bs = ad_date_to_bs_str(selected_date)
        except Exception:
            pass

    if not selected_date:
        selected_date = datetime.date.today()
        selected_date_bs = ad_date_to_bs_str(selected_date)

    # Invoices issued on this date
    invoices = Invoice.objects.filter(
        company=user_company,
        transaction_date=selected_date,
        is_deleted=False,
    ).exclude(status='CANCELLED').select_related('customer').order_by('created_at')

    # Payments received on this date
    payments = Payment.objects.filter(
        company=user_company,
        date=selected_date,
        is_deleted=False,
    ).select_related('invoice', 'invoice__customer').order_by('created_at')

    # Journal entries posted on this date
    journal_entries = JournalEntry.objects.filter(
        company=user_company,
        date=selected_date,
        is_deleted=False,
    ).prefetch_related('lines__account').order_by('created_at')

    # Totals
    total_invoiced = invoices.aggregate(t=Sum('total'))['t'] or Decimal('0')
    total_collected = payments.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    total_debits = Decimal('0')
    total_credits = Decimal('0')
    for je in journal_entries:
        for line in je.lines.all():
            if line.entry_type == 'DEBIT':
                total_debits += line.amount
            else:
                total_credits += line.amount

    context = {
        'selected_date': selected_date,
        'selected_date_bs': selected_date_bs,
        'invoices': invoices,
        'payments': payments,
        'journal_entries': journal_entries,
        'total_invoiced': total_invoiced,
        'total_collected': total_collected,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'rupee': 'रु',
    }
    return render(request, 'reports/bookkeeping/daybook_report.html', context)


@login_required
def export_daybook_report_excel(request):
    user_company = request.user_company
    if not user_company:
        messages.warning(request, "Your account is not associated with a company.")
        return redirect('accounts:user_dashboard')

    from apps.utils.nepali_date import bs_str_to_ad, ad_date_to_bs_str
    import datetime
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    date_bs_str = request.GET.get('date_bs', '').strip()
    date_ad_str = request.GET.get('date_ad', '').strip()
    selected_date = None

    if date_bs_str:
        try:
            selected_date = bs_str_to_ad(date_bs_str)
        except Exception:
            pass
    elif date_ad_str:
        try:
            selected_date = datetime.date.fromisoformat(date_ad_str)
        except Exception:
            pass
    if not selected_date:
        selected_date = datetime.date.today()

    selected_date_bs = ad_date_to_bs_str(selected_date)

    invoices = Invoice.objects.filter(
        company=user_company, transaction_date=selected_date, is_deleted=False,
    ).exclude(status='CANCELLED').select_related('customer')

    payments = Payment.objects.filter(
        company=user_company, date=selected_date, is_deleted=False,
    ).select_related('invoice__customer')

    journal_entries = JournalEntry.objects.filter(
        company=user_company, date=selected_date, is_deleted=False,
    ).prefetch_related('lines__account')

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')

    def _make_sheet(ws, title, headers):
        ws.title = title
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # Invoices sheet
    ws1 = wb.active
    _make_sheet(ws1, 'Invoices', ['Invoice #', 'Customer', 'Status', 'Total (NPR)', 'Outstanding (NPR)'])
    for row, inv in enumerate(invoices, 2):
        ws1.append([
            inv.invoice_number or str(inv.id),
            inv.customer.name if inv.customer else '—',
            inv.status,
            float(inv.total),
            float(inv.outstanding_balance),
        ])

    # Payments sheet
    ws2 = wb.create_sheet('Payments')
    _make_sheet(ws2, 'Payments', ['Invoice #', 'Customer', 'Amount (NPR)', 'Method'])
    for row, pay in enumerate(payments, 2):
        ws2.append([
            pay.invoice.invoice_number if pay.invoice else '—',
            pay.invoice.customer.name if pay.invoice and pay.invoice.customer else '—',
            float(pay.amount),
            pay.get_method_display() if pay.method else '—',
        ])

    # Journal entries sheet
    ws3 = wb.create_sheet('Journal Entries')
    _make_sheet(ws3, 'Journal Entries', ['Description', 'Account', 'Type', 'Amount (NPR)'])
    for je in journal_entries:
        for line in je.lines.all():
            ws3.append([
                je.description,
                line.account.name if line.account else '—',
                line.entry_type,
                float(line.amount),
            ])

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            ws.column_dimensions[letter].width = max(
                (len(str(c.value)) for c in col if c.value), default=10
            ) + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="daybook_{selected_date}.xlsx"'
    wb.save(response)
    return response
